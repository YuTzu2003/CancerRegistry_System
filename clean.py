import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from modules.rules import RULES
from modules.validate import check_error_type, validate_cell
from field_mapping import field_mapping

def cleanValidate(input_file, sheet_name, output_file):
    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
    alias_mapping, _ = field_mapping('中文欄位名稱')
    error_mask = pd.DataFrame("", index=df.index, columns=df.columns) 
    
    for col in df.columns:
        clean_col = str(col).strip()
        rule_name = alias_mapping.get(clean_col)

        if rule_name and rule_name in RULES:
            rule = RULES[rule_name]
            error_mask[col] = df[col].apply(lambda x: check_error_type(x, rule))
        elif clean_col in RULES:
            rule = RULES[clean_col]
            error_mask[col] = df[col].apply(lambda x: check_error_type(x, rule))

    df['_has_error'] = (error_mask != "").any(axis=1)
    
    # 將錯誤資料排到最上面
    sorted_df = df.sort_values(by='_has_error', ascending=False, kind='mergesort')
    sorted_mask = error_mask.loc[sorted_df.index]
    sorted_df = sorted_df.drop(columns=['_has_error'])
    sorted_df.to_excel(output_file, index=False, engine='openpyxl')

    # 錯誤標記
    wb = load_workbook(output_file)
    ws = wb.active
    fill_missing = PatternFill("solid", fgColor="FFE153")# 黃底=遺漏
    fill_format = PatternFill("solid", fgColor="FF9797") # 紅底=格式錯
    
    for r_idx in range(len(sorted_df)):
        for c_idx in range(len(sorted_df.columns)):
            err_type = sorted_mask.iloc[r_idx, c_idx]
            if err_type == "missing":
                ws.cell(row=r_idx + 2, column=c_idx + 1).fill = fill_missing
            elif err_type == "format":
                ws.cell(row=r_idx + 2, column=c_idx + 1).fill = fill_format

    wb.save(output_file)
    
    total_count = len(df)
    error_mask_bool = (error_mask != "")
    error_count = error_mask_bool.any(axis=1).sum()
    accuracy = (total_count - error_count) / total_count if total_count > 0 else 0
    
    # 計算錯誤類別比率
    missing_count = (error_mask == "missing").sum().sum()
    format_count = (error_mask == "format").sum().sum()
    
    stats = {
        'total': total_count,
        'errors': error_count,
        'accuracy': accuracy,
        'missing': missing_count,
        'format': format_count,
        'error_details': error_mask # 傳回完整的錯誤遮罩供後續顯示使用
    }

    return stats, alias_mapping, sorted_df, sorted_mask

if __name__ == "__main__": 
    input_file = 'data/20260318測試.xlsx'
    sheet_name = '1150318虛擬V1(給虎科)'
    output_file = f"validate_{sheet_name}.xlsx"
    cleanValidate(input_file, sheet_name, output_file)
