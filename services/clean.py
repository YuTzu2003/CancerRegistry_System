import os
import io
import csv
import uuid
import json
import shutil
import logging
import zipfile
import re
import pandas as pd
from copy import copy
from modules.db import get_conn
from contextlib import redirect_stdout
from modules.cleaner import cleanValidate
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from services.auth import login_required, admin_required
from modules.field_mapping import detect_system, get_field_map
from modules.clean_pipeline.validate import validate_date_rules
from flask import Blueprint, render_template, request, session, jsonify, send_file

clean_bp = Blueprint('clean', __name__)
Jobs_FOLDER = 'static/Jobs'


def _natural_sort_key(s):
    """
    Generate a key for natural sorting.
    Example: "7.10其他因子10" -> ['', 7, '.', 10, '其他因子10']
    """
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', str(s))]


def run_clean_validate_with_clean_log(*args, **kwargs):

    buffer = io.StringIO()

    with redirect_stdout(buffer):
        result = cleanValidate(*args, **kwargs)

    printed_text = buffer.getvalue().strip()

    for line in printed_text.splitlines():
        if line.startswith("Data Clean Report file:"):
            report_path = line.replace("Data Clean Report file:", "").strip()
            logging.info(f"Data Clean Report file created: {os.path.basename(report_path)}")
        elif line.strip():
            logging.info(line.strip())

    return result


DATE_ERROR_LIMIT = 3


def _job_files(project_path, original_filename, fmt_name):
    base_name, _ = os.path.splitext(original_filename)

    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")
    report_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Report.xlsx")
    working_file = os.path.join(project_path, f"{base_name}_working.xlsx")
    date_error_file = os.path.join(project_path, "date_errors.json")

    return base_name, cleaned_file, report_file, working_file, date_error_file


def _create_working_file(source_file, working_file):
    if os.path.abspath(source_file) != os.path.abspath(working_file):
        shutil.copy2(source_file, working_file)


def _build_date_errors(sorted_df, sorted_mask, alias_mapping, date_error_file):
    errors = []

    for row_pos, (source_row_index, mask_row) in enumerate(sorted_mask.iterrows()):
        date_cols = [
            col for col, val in mask_row.items()
            if val == "dateformat"
        ]

        if not date_cols:
            continue

        row_data = sorted_df.iloc[row_pos]
        _, msgs = validate_date_rules(row_data, alias_mapping)

        errors.append({
            "row_index": row_pos,
            "source_row_index": int(source_row_index),
            "excel_row": int(source_row_index) + 2,
            "fields": [
                {
                    "name": col,
                    "value": "" if pd.isna(row_data.get(col, "")) else str(row_data.get(col, ""))
                }
                for col in date_cols
            ],
            "messages": msgs or ["日期邏輯錯誤"]
        })

    with open(date_error_file, "w", encoding="utf-8") as f:
        json.dump(errors, f, ensure_ascii=False, indent=2)

    return errors


def _resolve_source_row_index(date_error_file, row_index):
    if not os.path.exists(date_error_file):
        return row_index

    try:
        with open(date_error_file, "r", encoding="utf-8") as f:
            date_errors = json.load(f)
    except (json.JSONDecodeError, OSError):
        return row_index

    for item in date_errors:
        if int(item.get("row_index", -1)) == int(row_index):
            return int(item.get("source_row_index", row_index))

    return row_index


def _load_job(job_id):
    conn = get_conn()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            Job.Path,
            Job.FileName,
            DataFormat.FmtName,
            DataFormat.Version,
            DataFormat.Revision_date
        FROM [Job]
        JOIN [DataFormat]
            ON Job.FmtID = DataFormat.FmtID
        WHERE Job.JobID = ?
    """, (job_id,))

    row = cursor.fetchone()
    conn.close()

    return row


def _update_working_file_cell_only(working_file, row_index, updates):
    wb = load_workbook(working_file)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_idx

    excel_row = int(row_index) + 2

    for field_name, new_value in updates.items():
        if field_name not in headers:
            continue

        col_idx = headers[field_name]
        cell = ws.cell(row=excel_row, column=col_idx)

        cell.value = (
            "" if new_value is None
            else str(new_value).replace("-", "").replace("/", "").strip()
        )

        # 只把被修改的日期欄位設成文字格式
        cell.number_format = "@"

    wb.save(working_file)
    wb.close()


def _detect_text_file_format(file_path):
    encodings = ["utf-8-sig", "utf-8", "cp950", "big5"]
    content = None
    used_encoding = "utf-8-sig"

    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                content = f.read(4096)
            used_encoding = enc
            break
        except UnicodeDecodeError:
            continue

    if content is None:
        content = ""

    first_line = content.splitlines()[0] if content.splitlines() else ""

    if "\t" in first_line:
        delimiter = "\t"
    elif "," in first_line:
        delimiter = ","
    elif ";" in first_line:
        delimiter = ";"
    else:
        ext = os.path.splitext(file_path)[1].lower()
        delimiter = "," if ext == ".csv" else "\t"

    return used_encoding, delimiter


def _cell_text(value):
    if value is None:
        return ""

    text = str(value).strip()

    if text.endswith(".0"):
        text = text[:-2]

    return text


def _working_file_rows(working_file):
    wb = load_workbook(working_file, data_only=False)
    ws = wb.active

    rows = [
        [_cell_text(value) for value in row]
        for row in ws.iter_rows(values_only=True)
    ]

    wb.close()
    return rows


def _truncate_to_encoded_width(text, width, encoding):
    result = ""

    for ch in text:
        candidate = result + ch
        if len(candidate.encode(encoding, errors="ignore")) > width:
            break
        result = candidate

    return result


def _format_fixed_width_value(value, width, encoding):
    text = _truncate_to_encoded_width(_cell_text(value), width, encoding)
    byte_len = len(text.encode(encoding, errors="ignore"))
    return text + (" " * max(width - byte_len, 0))


def _write_fixed_width_txt_from_working(working_file, source_file, fmt_name, encoding):
    fmt_val = str(fmt_name).replace("fmt_", "")
    field_spec = load_field_spec(fmt_val)
    rows = _working_file_rows(working_file)

    if not rows:
        return

    headers = rows[0]
    header_index = {name: idx for idx, name in enumerate(headers)}
    output_lines = []

    for row in rows[1:]:
        line_parts = []

        for field_name, start, end in field_spec:
            col_idx = header_index.get(field_name)
            value = row[col_idx] if col_idx is not None and col_idx < len(row) else ""
            line_parts.append(
                _format_fixed_width_value(value, end - start + 1, encoding)
            )

        output_lines.append("".join(line_parts))

    with open(source_file, "w", encoding=encoding, newline="") as f:
        f.write("\n".join(output_lines))
        if output_lines:
            f.write("\n")


def _write_csv_from_working(working_file, source_file, encoding, delimiter):
    rows = _working_file_rows(working_file)

    with open(source_file, "w", encoding=encoding, newline="") as f:
        writer = csv.writer(
            f,
            delimiter=delimiter,
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writerows(rows)


def _sync_working_file_to_source_files(working_file, source_file, converted_file, fmt_name):
    if not os.path.exists(working_file):
        return

    working_abs = os.path.abspath(working_file)

    if converted_file and os.path.exists(converted_file):
        converted_abs = os.path.abspath(converted_file)
        if converted_abs != working_abs:
            shutil.copy2(working_file, converted_file)

    if not source_file or not os.path.exists(source_file):
        return

    source_abs = os.path.abspath(source_file)
    ext = os.path.splitext(source_file)[1].lower()

    if ext in [".xlsx", ".xlsm"] and source_abs != working_abs:
        shutil.copy2(working_file, source_file)
        return

    if ext not in [".txt", ".csv"]:
        return

    encoding, delimiter = _detect_text_file_format(source_file)

    if ext == ".csv":
        _write_csv_from_working(working_file, source_file, encoding, delimiter)
        return

    _write_fixed_width_txt_from_working(
        working_file,
        source_file,
        fmt_name,
        encoding,
    )


def _build_cleaning_analysis(stats, sorted_mask):
    by_field = []

    for col in sorted_mask.columns:
        err_count = (sorted_mask[col] != "").sum()

        if err_count > 0:
            by_field.append({
                "name": col,
                "format": "檢核不符",
                "errors": int(err_count)
            })

    total_cells = stats["total"] * len(sorted_mask.columns)

    by_type = [
        {
            "type": "A:遺漏值",
            "count": int(stats["missing_cells"]),
            "ratio": f"{(stats['missing_cells'] / total_cells * 100):.1f}%" if total_cells > 0 else "0%"
        },
        {
            "type": "B:格式不符",
            "count": int(stats["format_cells"]),
            "ratio": f"{(stats['format_cells'] / total_cells * 100):.1f}%" if total_cells > 0 else "0%"
        },
        {
            "type": "C:邏輯錯誤",
            "count": int(stats["logic_cells"]),
            "ratio": f"{(stats['logic_cells'] / total_cells * 100):.1f}%" if total_cells > 0 else "0%"
        }
    ]

    return {
        "by_field": by_field,
        "by_type": by_type
    }


def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

@clean_bp.route("/api/categorize_fields", methods=["POST"])
@login_required
def api_categorize_fields():
    data = request.json
    job_id = data.get("job_id")
    scheme = data.get("scheme")

    if not job_id or not scheme:
        return jsonify({"ok": False, "error": "參數不足"}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("""SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID WHERE Job.JobID=?""", (job_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": False, "error": "找不到該紀錄"}), 404

    project_path, original_filename, fmt_name = row
    base_name, _ = os.path.splitext(original_filename)
    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")

    if not os.path.exists(cleaned_file):
        return jsonify({"ok": False, "error": "清洗檔案不存在"}), 404

    try:
        wb = load_workbook(cleaned_file, read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(max_row=1))]
        wb.close()

        # 取得當前格式的規則
        from modules.cleaner import FORMAT_RULES_MAP
        fmt_key = f"fmt_{fmt_name}" if not str(fmt_name).startswith("fmt_") else str(fmt_name)
        rules = FORMAT_RULES_MAP.get(fmt_key, {})
        id_to_rule_name = {v.get("ID"): k for k, v in rules.items() if v.get("ID")}

        alias_to_target = get_field_map(scheme, fmt_name)
        mapped = []
        unmapped = []
        for h in headers:
            if not h or h.startswith('_') or h == '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)':
                continue
            if h in alias_to_target:
                target_name = alias_to_target[h]
                final_display = target_name
                
                # 智慧換名：針對特定欄位進行精準二選一匹配
                if (scheme == 'field_name_zh' or scheme == '中文欄位名稱') and '/' in target_name:
                    m_target = re.match(r'^(\d+(\.\d+)*)', target_name)
                    if m_target:
                        seq = m_target.group(1)
                        if seq in ['4.2.1.8', '7.6']:
                            # 優先：如果來源欄位(h)剛好是其中一個拆分後的名稱，則採用該名稱
                            target_raw = target_name[len(seq):].strip()
                            valid_parts = [p.strip() for p in target_raw.split('/') if p.strip()]
                            source_raw = h[len(seq):].strip()
                            
                            if h.startswith(seq) and source_raw in valid_parts:
                                final_display = h
                            else:
                                # 次之：如果不是，則嘗試從 Python 規則中抓取該格式預設的中文名稱
                                rule_name = id_to_rule_name.get(seq)
                                if rule_name:
                                    final_display = f"{seq}{rule_name}"
                
                mapped.append({"key": h, "label": final_display, "target": final_display})
            else:
                unmapped.append({"key": h, "label": h})
        
        # 自然排序：確保 1.1, 1.2, ..., 1.10 的順序正確
        mapped.sort(key=lambda x: _natural_sort_key(x['label']))
        unmapped.sort(key=lambda x: _natural_sort_key(x['label']))

        return jsonify({"ok": True,"mapped": mapped,"unmapped": unmapped})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@clean_bp.route("/api/export", methods=["POST"])
@login_required
def api_export():
    data = request.json
    job_id = data.get("job_id")
    scheme = data.get("scheme")
    selected_fields = data.get("fields", [])
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("""SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID WHERE Job.JobID=?""", (job_id,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return jsonify({"ok": False, "error": "找不到該紀錄"}), 404

    project_path, original_filename, fmt_name = row
    base_name, _ = os.path.splitext(original_filename)
    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")

    if not os.path.exists(cleaned_file):
        return jsonify({"ok": False, "error": "清洗檔案不存在"}), 404

    alias_to_target = get_field_map(scheme, fmt_name)
    scheme_display = {
        "field_name_zh":"中文欄位名稱",
        "field_name_en":"英文欄位名稱",
        "ntu_yunlin":"台大雲林欄位名稱",
        "ntu_system":"台大體系醫整庫欄位名稱",
        "taiwan_cancer_registry":"台灣癌症登記中心",
        "AI_module":"雲醫癌AI模組"}.get(scheme, scheme)

    try:
        wb = load_workbook(cleaned_file)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        unique_targets = sorted(list(set(alias_to_target.values())), key=_natural_sort_key)
        output_cols = []
        used_orig_indices = set()

        # 1. ALWAYS include ALL fields defined for this module
        for target_name in unique_targets:
            source_idx = None
            final_display_name = target_name
            
            for i, h in enumerate(headers):
                if h in alias_to_target and alias_to_target[h] == target_name:
                    source_idx = i + 1
                    
                    # 智慧對應：僅針對特定欄位 (4.2.1.8, 7.6) 進行斜線拆分精準匹配
                    if scheme_display == '中文欄位名稱' and '/' in target_name:
                        m_target = re.match(r'^(\d+(\.\d+)*)', target_name)
                        if m_target:
                            seq_prefix = m_target.group(1)
                            if seq_prefix in ['4.2.1.8', '7.6']:
                                # 拆分資料庫全名為清單
                                target_raw_name = target_name[len(seq_prefix):].strip()
                                valid_parts = [p.strip() for p in target_raw_name.split('/') if p.strip()]
                                
                                # 取得來源欄位名稱部分
                                if h.startswith(seq_prefix):
                                    source_raw_name = h[len(seq_prefix):].strip()
                                    # 精準比對：必須等於左邊或右邊的其中一個
                                    if source_raw_name in valid_parts:
                                        final_display_name = h
                    break
            
            if source_idx:
                output_cols.append((source_idx, final_display_name))
                used_orig_indices.add(source_idx)
            else:
                output_cols.append((None, final_display_name))

        # 2. Add extra manually selected fields (that are NOT in the module)
        for field in selected_fields:
            if field in headers:
                orig_idx = headers.index(field) + 1
                if orig_idx not in used_orig_indices:
                    output_cols.append((orig_idx, field))
                    used_orig_indices.add(orig_idx)

        err_col_name = '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)'
        if err_col_name in headers:
            err_idx = headers.index(err_col_name) + 1
            if err_idx not in used_orig_indices:
                output_cols.append((err_idx, err_col_name))


        new_wb = Workbook()
        new_ws = new_wb.active
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            for c_idx, (orig_col_idx, display_name) in enumerate(output_cols, start=1):
                target_cell = new_ws.cell(row=r_idx, column=c_idx)
                
                if orig_col_idx is not None:
                    source_cell = ws.cell(row=r_idx, column=orig_col_idx)
                    if r_idx == 1:
                        target_cell.value = display_name
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.alignment = copy(source_cell.alignment)
                    else:
                        copy_cell(source_cell, target_cell)
                else:
                    if r_idx == 1:
                        target_cell.value = display_name
                    else:
                        target_cell.value = ""

        for c_idx, (orig_col_idx, _) in enumerate(output_cols, start=1):
            if orig_col_idx is not None:
                new_ws.column_dimensions[get_column_letter(c_idx)].width = \
                    ws.column_dimensions[get_column_letter(orig_col_idx)].width
            else:
                new_ws.column_dimensions[get_column_letter(c_idx)].width = 15

        export_filename = f"fmt{fmt_name}_{base_name}_{scheme_display}.xlsx"
        temp_out = os.path.join(project_path, export_filename)
        new_wb.save(temp_out)
        resp = send_file(os.path.abspath(temp_out), as_attachment=True, download_name=export_filename)
        resp.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@clean_bp.route("/api/preview", methods=["POST"])
@login_required
def api_preview():
    data = request.json
    job_id = data.get("job_id")
    scheme = data.get("scheme")
    selected_fields = data.get("fields", [])
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("""SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] 
                    JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID 
                    WHERE Job.JobID=?""", (job_id,))
    row = cursor.fetchone()
    conn.close()

    project_path, original_filename, fmt_name = row
    base_name, _ = os.path.splitext(original_filename)
    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")

    if not os.path.exists(cleaned_file):
        return jsonify({"ok": False, "error": "清洗檔案不存在"}), 404
    alias_to_target = get_field_map(scheme, fmt_name)

    try:
        wb = load_workbook(cleaned_file, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        unique_targets = sorted(list(set(alias_to_target.values())), key=_natural_sort_key)
        output_cols = []
        used_orig_indices = set()

        # 1. ALWAYS include ALL fields defined for this module
        for target_name in unique_targets:
            source_idx = None
            final_display_name = target_name
            
            for i, h in enumerate(headers):
                if h in alias_to_target and alias_to_target[h] == target_name:
                    source_idx = i + 1
                    
                    # 智慧對應：僅針對特定欄位 (4.2.1.8, 7.6) 進行斜線拆分精準匹配
                    if (scheme == 'field_name_zh' or scheme == '中文欄位名稱') and '/' in target_name:
                        m_target = re.match(r'^(\d+(\.\d+)*)', target_name)
                        if m_target:
                            seq_prefix = m_target.group(1)
                            if seq_prefix in ['4.2.1.8', '7.6']:
                                target_raw_name = target_name[len(seq_prefix):].strip()
                                valid_parts = [p.strip() for p in target_raw_name.split('/') if p.strip()]
                                if h.startswith(seq_prefix):
                                    source_raw_name = h[len(seq_prefix):].strip()
                                    if source_raw_name in valid_parts:
                                        final_display_name = h
                    break
            
            if source_idx:
                output_cols.append((source_idx, final_display_name))
                used_orig_indices.add(source_idx)
            else:
                output_cols.append((None, final_display_name))
        
        # 2. Add extra manually selected fields
        for f in selected_fields:
            if f in headers:
                idx = headers.index(f) + 1
                if idx not in used_orig_indices:
                    output_cols.append((idx, f))
                    used_orig_indices.add(idx)

        err_f = '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)'
        if err_f in headers:
            idx = headers.index(err_f) + 1
            if idx not in used_orig_indices:
                output_cols.append((idx, err_f))

        preview_data = []
        preview_headers = [col[1] for col in output_cols]
        for r_idx in range(2, 7):
            row_data = []
            for c_idx, _ in output_cols:
                if c_idx is not None:
                    val = ws.cell(row=r_idx, column=c_idx).value
                    row_data.append(str(val or ""))
                else:
                    row_data.append("")
            preview_data.append(row_data)
        return jsonify({"ok": True,"headers": preview_headers,"data": preview_data})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

def load_field_spec(fmt_val):
    conn = get_conn()
    cursor = conn.cursor()
    
    # 1. 取得該格式的所有欄位定義
    sql_fields = """SELECT [ChineseName], [Start], [End] FROM [CancerRegistry_Fields] 
                   WHERE [fmt]=? ORDER BY [Start]"""
    cursor.execute(sql_fields, fmt_val)
    field_rows = cursor.fetchall()
    
    # 2. 取得所有的 Mapping 資料，建立 名稱 -> 序號 的對照字典
    sql_map = "SELECT [序號], [中文欄位名稱] FROM [CancerRegistry_FieldMap]"
    cursor.execute(sql_map)
    map_rows = cursor.fetchall()
    conn.close()
    
    # 建立一個智慧對照字典：包含完整名稱以及拆解後的名稱
    name_to_seq = {}
    for r_seq, r_name in map_rows:
        if not r_seq or not r_name: continue
        seq_str = str(r_seq).strip().replace('.0', '')
        full_name = str(r_name).strip()
        
        # 加入完整名稱
        name_to_seq[full_name] = seq_str
        
        # 針對特定欄位處理斜線拆解
        if seq_str in ['4.2.1.8', '7.6'] and '/' in full_name:
            for part in full_name.split('/'):
                part = part.strip()
                if part:
                    name_to_seq[part] = seq_str

    # 3. 取得 Python 規則定義
    from modules.cleaner import FORMAT_RULES_MAP
    fmt_key = f"fmt_{fmt_val}"
    rules = FORMAT_RULES_MAP.get(fmt_key, {})
    id_to_rule_name = {r_val.get('ID'): r_key for r_key, r_val in rules.items() if r_val.get('ID')}
    
    field_spec = []
    for f_name, f_start, f_end in field_rows:
        orig_name = str(f_name).strip()
        
        # 智慧尋找序號
        seq = name_to_seq.get(orig_name, "")
        
        # 智慧決定顯示名稱
        target_name = orig_name
        if seq in ['4.2.1.8', '7.6']:
            rule_name = id_to_rule_name.get(seq)
            if rule_name:
                # 如果 Python 規則名稱存在於 Mapping 的中文名稱中，則採用 Python 的簡短名稱
                # 這裡需要找到對應的 Full Name 進行比對
                # 為了效能與精準度，我們直接判斷 rule_name 是否在剛才配對到的 seq 對應的 full_name 裡
                for m_seq, m_full in map_rows:
                    if str(m_seq).strip().replace('.0', '') == seq:
                        if rule_name in str(m_full):
                            target_name = rule_name
                            break

        header_name = f"{seq}{target_name}" if seq else target_name
        field_spec.append((header_name, int(f_start), int(f_end)))
        
    return field_spec

def parse_fixed_width_line(line_text, spec):
    line_bytes = line_text.encode('big5', errors='ignore')
    parsed_row = {}
    for name, start, end in spec:
        field_value = line_bytes[start - 1:end]
        try:
            parsed_row[name] = field_value.decode('big5').strip()
        except:
            parsed_row[name] = field_value.decode('big5', errors='replace').strip()
    return parsed_row

@clean_bp.route("/clean")
@login_required
def clean():
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT [FmtID], [FmtName], [Version], [Revision_date] FROM [DataFormat] ORDER BY FmtName ASC")
    formats = [{"id": str(r.FmtID), "name": str(r.FmtName), "version": str(r.Version), "updated": str(r.Revision_date)} for r in cursor.fetchall()]
    conn.close()
    return render_template("clean.html", active="clean", formats=formats)

@clean_bp.route("/api/formats", methods=["POST"])
@admin_required
def api_add_format():
    if request.is_json:
        data = request.json
    else:
        data = request.form
    name, version, updated = data.get("name"), data.get("version"), data.get("updated")
    if not name or not version: return jsonify({"ok": False, "message": "名稱與版本為必填"}), 400
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO [DataFormat] ([FmtName], [Version], [Revision_date]) VALUES (?, ?, ?)", (name, version, updated))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@clean_bp.route("/api/formats/<fmt_id>", methods=["PUT", "DELETE", "POST"])
@admin_required
def api_manage_format(fmt_id):
    conn = get_conn()
    cursor = conn.cursor()

    if request.method == "DELETE":
        cursor.execute("DELETE FROM [DataFormat] WHERE [FmtID] = ?", (fmt_id,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    if request.is_json:
        data = request.json
    else:
        data = request.form 
    name, version, updated = data.get("name"), data.get("version"), data.get("updated")
    cursor.execute("UPDATE [DataFormat] SET [FmtName]=?, [Version]=?, [Revision_date]=? WHERE [FmtID]=?", (name, version, updated, fmt_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@clean_bp.route("/api/cleanJob", methods=["POST"])
@login_required
def api_clean():
    user_id, format_id = session.get("id"), request.form.get("format_id")
    uploaded_file = request.files.get("data_file")
    if not format_id or not uploaded_file or uploaded_file.filename == '': return jsonify({"ok": False, "error": "未選擇檔案"}), 400
    JobID = str(uuid.uuid4())
    project_folder = f"{Jobs_FOLDER}/{JobID}"
    os.makedirs(project_folder, exist_ok=True)

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT [FmtName], [Version], [Revision_date] FROM [DataFormat] WHERE [FmtID] = ?", (format_id,))
    fmt_data = cursor.fetchone()
    if not fmt_data:
        conn.close()
        return jsonify({"ok": False, "error": "找不到指定的格式"}), 400
        
    fmt_name, version, rev_date = fmt_data
    filename = os.path.basename(uploaded_file.filename)
    if not filename:
        filename = "uploaded_file"
        
    path = f"{project_folder}/{filename}"
    uploaded_file.save(path)
    uploaded_file.close()
    
    file_ext = os.path.splitext(filename)[1].lower()
    base_name = os.path.splitext(filename)[0]
    process_path = path

    if file_ext == '.csv':
        try:
            df_csv = pd.read_csv(path, dtype=str, encoding='utf-8-sig')
        except UnicodeDecodeError:
            df_csv = pd.read_csv(path, dtype=str, encoding='cp950')
        
        temp_xlsx = f"{project_folder}/{base_name}.xlsx"
        df_csv.to_excel(temp_xlsx, index=False)
        process_path = temp_xlsx

    if file_ext == '.txt':
        try:
            fmt_val = str(fmt_name).replace("fmt_", "")
            cursor.execute("SELECT MAX([End]) FROM CancerRegistry_Fields WHERE [fmt]=? GROUP BY [fmt]", (fmt_val,))
            row = cursor.fetchone()
            expected_length = row[0] if row else 0

            length_errors = []
            # 1. 預檢長度
            with open(path, 'r', encoding='big5', errors='ignore') as f:
                for i, line in enumerate(f):
                    clean_line = line.rstrip('\n').rstrip('\r')
                    if not clean_line.strip(): continue
                    length = len(clean_line.encode('big5', errors='ignore'))
                    if expected_length > 0 and length != expected_length:
                        length_errors.append(f"第 {i+1} 行: 實際 {length},預期 {expected_length}")
                        
            # 整理錯誤訊息，避免畫面過長
            error_message_html = ""
            if length_errors:
                total_errors = len(length_errors)
                display_limit = 3

                error_message_html += f"<b>檔案長度校驗失敗，總共發現 {total_errors} 筆長度不符的資料：</b><br>"

                # 顯示前 N 筆
                error_message_html += "<br>".join(length_errors[:display_limit])
                
                if total_errors > display_limit:
                    error_message_html += f"<br>...以及其他 {total_errors - display_limit} 筆錯誤。"

            # 2. 不論長度是否正確，都進行初步切分並產生 Excel (供下載檢視)
            field_spec = load_field_spec(fmt_val)
            results = []
            with open(path, 'r', encoding='big5', errors='ignore') as f:
                for line in f:
                    # 即使該行長度有錯，parse_fixed_width_line 也會盡力切分
                    results.append(parse_fixed_width_line(line, field_spec))

            temp_xlsx = f"{project_folder}/{base_name}.xlsx"
            keys = [f[0] for f in field_spec]
            wb_temp = Workbook()
            ws_temp = wb_temp.active
            ws_temp.append(keys)

            for row_dict in results:
                row_values = [row_dict.get(k, "") for k in keys]
                ws_temp.append(row_values)
            
            # 設定所有儲存格為文字格式，避免 Excel 自動轉換 (如 1.10 變 1.1)
            for row in ws_temp.iter_rows():
                for cell in row:
                    cell.number_format = '@'
            
            wb_temp.save(temp_xlsx)
            process_path = temp_xlsx

            # 3. 如果有長度錯誤，立即回傳並提供下載按鈕所需的資訊
            if length_errors:
                # 產生完整的錯誤 Log 檔
                log_path = f"{project_folder}/length_errors.log"
                with open(log_path, 'w', encoding='utf-8') as log_f:
                    log_f.write(f"檔案名稱: {filename}\n")
                    log_f.write("\n".join(length_errors))
                
                # 必須先在資料庫建立 Job 紀錄，下載 API 才能根據 JobID 找到路徑
                cursor.execute("INSERT INTO Job ([JobID],[UserID],[FmtID],[FileName],[TotalCount],[CompletenessScore],[CorrectScore],[ConsistencyScore],[DQI],[Path]) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                (JobID, user_id, format_id, filename, 0, 0, 0, 0, 0, project_folder))
                conn.commit()
                conn.close()
                return jsonify({
                    "ok": False, 
                    "error": error_message_html, 
                    "has_length_error": True,
                    "job_id": JobID
                }), 400

        except Exception as e:
            if conn and not conn.closed: conn.close()
            return jsonify({"ok": False, "error": f"TXT 解析失敗: {str(e)}"}), 500

    base_name, out_path, rep_path, working_file, date_error_file = _job_files(project_folder, filename, fmt_name)
    
    try:
        _create_working_file(process_path, working_file)

        stats, alias_mapping, sorted_df, sorted_mask = cleanValidate(process_path, out_path, rep_path, f"fmt_{fmt_name}", version, rev_date)

        date_errors = _build_date_errors(sorted_df, sorted_mask, alias_mapping, date_error_file)

        cursor.execute("INSERT INTO Job ([JobID],[UserID],[FmtID],[FileName],[TotalCount],[CompletenessScore],[CorrectScore],[ConsistencyScore],[DQI],[Path]) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (JobID, user_id, format_id, filename, int(stats['total']), float(stats['completeness']), float(stats['correctness']), float(stats['consistency']), float(stats['quality_score']), project_folder))
        conn.commit()
    except Exception as e:
        if conn and not conn.closed: conn.rollback()
        return jsonify({"ok": False, "error":str(e)}), 500
    finally: 
        if conn and not conn.closed:
            conn.close()

    by_field = []
    for col in sorted_mask.columns:
        err_count = (sorted_mask[col] != "").sum()
        if err_count > 0:
            by_field.append({"name": col,"format": "檢核不符","errors": int(err_count)})
    
    by_type = [
        {"type": "A:遺漏值", "count": int(stats['missing_cells']), "ratio": f"{(stats['missing_cells']/(stats['total']*len(sorted_mask.columns))*100):.1f}%" if stats['total'] > 0 else "0%"},
        {"type": "B:格式不符", "count": int(stats['format_cells']), "ratio": f"{(stats['format_cells']/(stats['total']*len(sorted_mask.columns))*100):.1f}%" if stats['total'] > 0 else "0%"},
        {"type": "C:邏輯錯誤", "count": int(stats['logic_cells']), "ratio": f"{(stats['logic_cells']/(stats['total']*len(sorted_mask.columns))*100):.1f}%" if stats['total'] > 0 else "0%"}
    ]

    output_fields = [{"key": col, "label": col} for col in sorted_df.columns if not col.startswith('_')]
    
    # 偵測資料體系
    detected_system, _ = detect_system(sorted_df.columns)

    date_error_count = len(date_errors)

    if date_error_count > DATE_ERROR_LIMIT:
        message = (
            f"日期邏輯錯誤共有 {date_error_count} 筆，"
            f"已達系統限制 {DATE_ERROR_LIMIT} 筆，"
            "請先修正錯誤的日期資料，完成修正後再進行後續資料清洗作業"
        )
        can_continue = False
    else:
        message = "清洗完成"
        can_continue = True

    return jsonify({
        "ok": True,
        "project_id": JobID,
        "job_id": JobID,
        "detected_system": detected_system,
        "message": message,
        "can_continue": can_continue,
        "stats": {
            "total": int(stats['total']), 
            "passed": int(stats['total'] - stats['error_rows']), 
            "error": int(stats['error_rows']), 
            "completeness": float(stats['completeness']), 
            "correctness": float(stats['correctness']), 
            "consistency": float(stats['consistency']),
            "dqi": float(stats['quality_score'])
        },
        "analysis": {"by_field": by_field,"by_type": by_type},
        "output_fields": output_fields,
        "date_error_limit": DATE_ERROR_LIMIT,
        "date_error_count": date_error_count,
        "date_errors": date_errors
    })


@clean_bp.route("/api/date_errors", methods=["POST"])
@login_required
def api_date_errors():
    data = request.get_json(silent=True) or {}
    job_id = data.get("job_id")

    if not job_id:
        return jsonify({"ok": False, "error": "缺少 job_id"}), 400

    row = _load_job(job_id)

    if not row:
        return jsonify({"ok": False, "error": "找不到該紀錄"}), 404

    project_path, original_filename, fmt_name, version, rev_date = row

    _, _, _, _, date_error_file = _job_files(project_path, original_filename, fmt_name)

    if not os.path.exists(date_error_file):
        return jsonify({
            "ok": True,
            "date_errors": [],
            "date_error_count": 0,
            "date_error_limit": DATE_ERROR_LIMIT
        })

    with open(date_error_file, "r", encoding="utf-8") as f:
        date_errors = json.load(f)

    return jsonify({
        "ok": True,
        "date_errors": date_errors,
        "date_error_count": len(date_errors),
        "date_error_limit": DATE_ERROR_LIMIT
    })


@clean_bp.route("/api/date_errors/update", methods=["POST"])
@login_required
def api_update_date_error():
    data = request.get_json(silent=True) or {}

    job_id = data.get("job_id")
    row_index = data.get("row_index")
    updates = data.get("updates") or {}

    if not job_id:
        return jsonify({"ok": False, "error": "缺少 job_id"}), 400

    if row_index is None or not isinstance(updates, dict) or not updates:
        return jsonify({"ok": False, "error": "缺少修正資料"}), 400

    row = _load_job(job_id)

    if not row:
        return jsonify({"ok": False, "error": "找不到該紀錄"}), 404

    project_path, original_filename, fmt_name, version, rev_date = row

    _, cleaned_file, report_file, working_file, date_error_file = _job_files(
        project_path,
        original_filename,
        fmt_name
    )

    if not os.path.exists(working_file):
        return jsonify({"ok": False, "error": "找不到可修正的暫存資料"}), 404

    try:
        row_index = int(row_index)
    except ValueError:
        return jsonify({"ok": False, "error": "修正列格式錯誤"}), 400

    try:
        source_row_index = _resolve_source_row_index(date_error_file, row_index)

        _update_working_file_cell_only(working_file, source_row_index, updates)

        source_file = os.path.join(project_path, original_filename)
        base_name, _ = os.path.splitext(original_filename)
        converted_file = os.path.join(project_path, f"{base_name}.xlsx")

        _sync_working_file_to_source_files(
            working_file,
            source_file,
            converted_file,
            fmt_name
        )

        stats, alias_mapping, sorted_df, sorted_mask = run_clean_validate_with_clean_log(
            working_file,
            cleaned_file,
            report_file,
            f"fmt_{fmt_name}",
            version,
            rev_date
        )

        date_errors = _build_date_errors(sorted_df, sorted_mask, alias_mapping, date_error_file)

        conn = get_conn()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE [Job]
            SET 
                TotalCount = ?,
                CompletenessScore = ?,
                CorrectScore = ?,
                ConsistencyScore = ?,
                DQI = ?
            WHERE JobID = ?
        """, (
            int(stats["total"]),
            float(stats["completeness"]),
            float(stats["correctness"]),
            float(stats["consistency"]),
            float(stats["quality_score"]),
            job_id
        ))

        conn.commit()
        conn.close()

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({
        "ok": True,
        "message": "修正完成，已重新檢核",
        "job_id": job_id,
        "project_id": job_id,
        "date_error_limit": DATE_ERROR_LIMIT,
        "date_error_count": len(date_errors),
        "date_errors": date_errors,
        "analysis": _build_cleaning_analysis(stats, sorted_mask),
        "stats": {
            "total": int(stats["total"]),
            "passed": int(stats["total"] - stats["error_rows"]),
            "error": int(stats["error_rows"]),
            "completeness": float(stats["completeness"]),
            "correctness": float(stats["correctness"]),
            "consistency": float(stats["consistency"]),
            "dqi": float(stats["quality_score"])
        }
    })

@clean_bp.route("/api/download/<file_type>/<job_id>")
@login_required
def download_file(file_type, job_id):
    try:
        conn = get_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID WHERE Job.JobID=?", (job_id,))
        row = cursor.fetchone()
        if not row: return jsonify({"ok": False, "error": "找不到該紀錄"}), 404
        project_path, original_filename, fmt_name = row
        base_name, _ = os.path.splitext(original_filename)
        
        display_name = None
        if file_type == "cleaned":
            target_filename = f"fmt{fmt_name}_{base_name}_Clean.xlsx"
        elif file_type == "report":
            target_filename = f"fmt{fmt_name}_{base_name}_Report.xlsx"
        elif file_type == "original":
            target_filename = original_filename
        elif file_type == "parsed" or file_type == "preview":
            target_filename = f"{base_name}.xlsx"
            display_name = f"欄位檢核表_{base_name}.xlsx"
        elif file_type == "length_log":
            target_filename = "length_errors.log"
            display_name = f"欄位檢核表_{base_name}.log"
        else:
            return jsonify({"ok": False, "error": "無效的下載類型"}), 400

        file_path = os.path.join(project_path, target_filename)
        conn.close()
        if not os.path.exists(file_path): 
            return jsonify({"ok": False, "error": "檔案不存在"}), 404
            
        return send_file(
            os.path.abspath(file_path), 
            as_attachment=True, 
            download_name=display_name if display_name else target_filename
        )
    except Exception as e: return jsonify({"ok": False, "error": str(e)}), 500
