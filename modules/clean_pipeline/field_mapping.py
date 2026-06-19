import pandas as pd
import re
from openpyxl import load_workbook
from modules.db import get_conn

def detect_system(excel_columns):
    conn = get_conn()
    systems = ['中文欄位名稱','英文欄位名稱','台大雲林欄位名稱','台大體系醫整庫欄位名稱','台灣癌症登記中心','雲醫癌AI模組']
    columns_sql = ', '.join(f'[{s}]' for s in systems)

    query = f"SELECT [序號], {columns_sql} FROM [Hospital_data].[dbo].[CancerRegistry_FieldMap]"
    cursor = conn.cursor()
    cursor.execute(query)
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]
    df_mapping = pd.DataFrame.from_records(rows, columns=columns)
    conn.close()

    # 去除空格並轉小寫，以便進行不分大小寫與空白的對照
    excel_cols_set = set(re.sub(r'\s+', '', str(col)).strip().lower() for col in excel_columns if col)
    scores = {}
    for s in systems:
        db_cols = set()
        for _, row in df_mapping.iterrows():
            seq = str(row['序號']).strip()
            if seq.endswith('.0'):
                seq = seq[:-2]
                
            val = row[s]
            if pd.notna(val):
                clean_val = str(val).strip()
                if clean_val:
                    if seq in ['4.2.1.8', '7.6'] and '/' in clean_val:
                        parts = [p.strip() for p in clean_val.split('/') if p.strip()]
                    else:
                        parts = [clean_val]
                        
                    for part in parts:
                        norm_part = re.sub(r'\s+', '', part).lower()
                        # 支援有序號前置的寫法 (如 1.2personid2)
                        db_cols.add(f"{seq.lower()}{norm_part}")
                        # 支援無序號前置的寫法 (如 personid2)
                        db_cols.add(norm_part)
                        
        match_count = len(excel_cols_set.intersection(db_cols))
        scores[s] = match_count

    detected_system = max(scores, key=scores.get)
    if scores[detected_system] == 0:
        return "unknown", scores
    return detected_system, scores

def field_mapping(target_col):
    conn = get_conn()
    query = """SELECT * FROM [Hospital_data].[dbo].[CancerRegistry_FieldMap]"""
    cursor = conn.cursor()
    cursor.execute(query)
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]
    df_mapping = pd.DataFrame.from_records(rows, columns=columns)
    conn.close()

    alias_dict = {}
    output_field_list = []
    for _, row in df_mapping.iterrows():
        val = row[target_col]
        output_name = str(val).strip() if pd.notna(val) else ""
        
        seq = str(row['序號']).strip()
        if seq.endswith('.0'):
            seq = seq[:-2]

        if output_name and output_name not in output_field_list:
            output_field_list.append(output_name)

        aliases = [
            row['中文欄位名稱'], 
            row['英文欄位名稱'], 
            row['台大雲林欄位名稱'], 
            row['台大體系醫整庫欄位名稱'], 
            row['台灣癌症登記中心']
        ]

        for alias in aliases:
            if pd.notna(alias):
                clean_alias = str(alias).strip()
                if clean_alias:
                    if seq in ['4.2.1.8', '7.6'] and '/' in clean_alias:
                        parts = [p.strip() for p in clean_alias.split('/') if p.strip()]
                    else:
                        parts = [clean_alias]

                    for part in parts:
                        final_target_val = output_name
                        if target_col == '中文欄位名稱' and '/' in output_name and seq in ['4.2.1.8', '7.6']:
                            output_parts = [op.strip() for op in output_name.split('/')]
                            if part in output_parts:
                                final_target_val = part
                        
                        alias_dict[f"{seq}{part}"] = final_target_val

    return alias_dict, output_field_list

def get_field_map(target_scheme_key, fmt_name):
    scheme_map = {
        "field_name_zh":"中文欄位名稱",
        "field_name_en":"英文欄位名稱",
        "ntu_yunlin":"台大雲林欄位名稱",
        "ntu_system":"台大體系醫整庫欄位名稱",
        "taiwan_cancer_registry":"台灣癌症登記中心",
        "AI_module":"雲醫癌AI模組"
    }
    
    target_col = scheme_map.get(target_scheme_key,"中文欄位名稱")
    conn = get_conn()
    clean_fmt = str(fmt_name).replace("fmt_", "")
    query = f"""SELECT [序號],[中文欄位名稱],[英文欄位名稱],[台大雲林欄位名稱],[台大體系醫整庫欄位名稱],[台灣癌症登記中心],[雲醫癌AI模組]
                FROM [Hospital_data].[dbo].[v_FieldMap_WithFmt]
                WHERE [{clean_fmt}欄位] = 1"""
    cursor = conn.cursor()
    cursor.execute(query)
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]
    df_mapping = pd.DataFrame.from_records(rows, columns=columns)
    conn.close()

    alias_to_target = {}
    for _, row in df_mapping.iterrows():
        target_base_name = str(row[target_col]).strip() if pd.notna(row[target_col]) else ""
        if not target_base_name:
            continue
        seq = str(row['序號']).strip()
        if seq.endswith('.0'):
            seq = seq[:-2]
        
        target_name = f"{seq}{target_base_name}"

        for col in columns:
            if col == '序號': continue
            val = row[col]
            if pd.notna(val):
                alias = str(val).strip()
                if alias:
                    if seq in ['4.2.1.8', '7.6'] and '/' in alias:
                        parts = [p.strip() for p in alias.split('/') if p.strip()]
                    else:
                        parts = [alias]
                        
                    for part in parts:
                        alias_to_target[f"{seq}{part}"] = target_name
                    
    return alias_to_target
                    
    return alias_to_target

def process_data(excel_path, mapping_dict, AImodule_list, target_sheet=0):
    df_excel = pd.read_excel(excel_path, sheet_name=target_sheet)
    rename_map = {}
    for col in df_excel.columns:
        clean_col = str(col).strip()
        ai_name = mapping_dict.get(clean_col, "")
        if ai_name:
            rename_map[col] = ai_name
            
    df_transformed = df_excel[list(rename_map.keys())].copy()
    df_transformed.rename(columns=rename_map, inplace=True)
    for ai_name in AImodule_list:
        if ai_name not in df_transformed.columns:
            df_transformed[ai_name] = "" 
    df_final = df_transformed[AImodule_list]
    return df_final

def validate_and_rename_headers(headers, fmt_name):
    """
    驗證並重新命名上傳檔案的標頭。
    規則：
    1. 不可只有序號（如 1.2），必須序號+欄位名稱。
    2. 欄位名稱必須與 SQL Server 對照表吻合。
    3. 只有欄位名稱、序號+欄位名稱、序號+空格+欄位名稱都是合規的。
    4. 輸出統一格式為「序號+標準中文欄位名稱」（無空格，如 1.2病歷號碼）。
    5. 4.2.1.8 與 7.6 欄位只要能與其中一邊匹配即可。
    6. 其他無關的自訂欄位，若無法匹配且非忽略欄位，則拋出詳細的 ValueError 錯誤。
    """
    
    from modules.clean_pipeline.cleaner import FORMAT_RULES_MAP
    norm_fmt = f"fmt_{str(fmt_name).replace('fmt_', '')}"
    rules = FORMAT_RULES_MAP.get(norm_fmt, {})

    seq_to_std_name = {}
    for rule_name, rule_val in rules.items():
        r_id = str(rule_val.get("ID", "")).strip()
        if r_id:
            seq_to_std_name[r_id] = rule_name

    conn = get_conn()
    clean_fmt = str(fmt_name).replace("fmt_", "")
    query = f"""
        SELECT [序號], [中文欄位名稱], [英文欄位名稱], [台大雲林欄位名稱], [台大體系醫整庫欄位名稱], [台灣癌症登記中心], [雲醫癌AI模組]
        FROM [Hospital_data].[dbo].[v_FieldMap_WithFmt]
        WHERE [{clean_fmt}欄位] = 1
    """
    cursor = conn.cursor()
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()

    seq_to_aliases = {}  # seq -> set of normalized aliases
    alias_to_seqs = {}   # normalized_alias -> list of (seq, std_name)

    for row in rows:
        seq = str(row[0]).strip()
        if seq.endswith('.0'):
            seq = seq[:-2]
        
        std_name = seq_to_std_name.get(seq)
        if not std_name:
            continue

        if seq not in seq_to_aliases:
            seq_to_aliases[seq] = set()

        aliases = []
        for val in row[1:]:
            if val is not None:
                aliases.append(str(val).strip())
        
        for alias in aliases:
            if not alias:
                continue
            if seq in ['4.2.1.8', '7.6'] and '/' in alias:
                parts = [p.strip() for p in alias.split('/') if p.strip()]
            else:
                parts = [alias]
            
            for part in parts:
                norm_part = re.sub(r'\s+', '', part)
                if norm_part:
                    seq_to_aliases[seq].add(norm_part)
                    if norm_part not in alias_to_seqs:
                        alias_to_seqs[norm_part] = []
                    if (seq, std_name) not in alias_to_seqs[norm_part]:
                        alias_to_seqs[norm_part].append((seq, std_name))

    renamed_headers = []
    ignored_headers = {
        '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)',
        '錯誤註記說明'
    }

    for col in headers:
        col_str = str(col).strip()
        if not col_str or col_str.startswith('_') or col_str in ignored_headers:
            renamed_headers.append(col)
            continue

        match = re.match(r"^(\d+(?:\.\d+)+)(.*)$", col_str)
        if match:
            seq = match.group(1).strip()
            # 以序號作為優先判定條件：只要序號存在於該格式中，就直接匹配為統一格式 {序號}{標準中文欄位名稱}
            if seq in seq_to_std_name:
                std_name = seq_to_std_name[seq]
                renamed_headers.append(f"{seq}{std_name}")
                continue

        clean_col = re.sub(r'\s+', '', col_str)
        if clean_col in alias_to_seqs:
            matches = alias_to_seqs[clean_col]
            if matches:
                seq, std_name = matches[0]
                renamed_headers.append(f"{seq}{std_name}")
                continue
        
        renamed_headers.append(col)

    seen = set()
    duplicates = []
    for h in renamed_headers:
        if h in ignored_headers or not h or str(h).startswith('_'):
            continue
        if h in seen:
            duplicates.append(h)
        seen.add(h)
    if duplicates:
        raise ValueError(f"欄位標頭錯誤：重複的欄位對應「{', '.join(duplicates)}」，請確認上傳的欄位是否重複。")

    return renamed_headers

def validate_and_unify_headers_in_file(file_path, fmt_name):
    """
    讀取 Excel 檔案的第一行標頭並執行驗證與統一重組，接著寫回原 Excel 檔案以保留原始樣式。
    """
    wb = load_workbook(file_path)
    ws = wb.active
    
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")
        
    new_headers = validate_and_rename_headers(headers, fmt_name)
    
    for idx, new_val in enumerate(new_headers, start=1):
        ws.cell(row=1, column=idx).value = new_val
        
    wb.save(file_path)
    wb.close()

if __name__ == "__main__": 
    data_path = "data/20260318測試.xlsx"  
    sheet_name = "測1.1" 
    target_column = "雲醫癌AI模組"

    alias_mapping, all_fields = field_mapping(target_column)
    df_output_data = process_data(data_path, alias_mapping, all_fields, target_sheet=sheet_name)
    
    rename_result_path = f"data/{target_column}_output.xlsx"
    df_output_data.to_excel(rename_result_path, index=False)