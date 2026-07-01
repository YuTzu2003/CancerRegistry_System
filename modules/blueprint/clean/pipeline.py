import os
import io
import csv
import uuid
import json
import shutil
import logging
import re
import base64
import pandas as pd
from copy import copy
from modules.services.db import get_conn
from contextlib import redirect_stdout
from modules.blueprint.clean.cleaner import cleanValidate
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from modules.blueprint.clean.field_mapping import detect_system, get_field_map, validate_and_rename_headers, validate_and_unify_headers_in_file
from modules.blueprint.clean.text_converter import convert_txt_to_excel
from modules.blueprint.clean.rules.validate import validate_date_rules

def _natural_sort_key(s):
    def convert(text, is_match):
        if is_match:
            return tuple(int(x) for x in text.split('.'))
        return text.lower()
    parts = re.split(r'(\d+(?:\.\d+)*)', str(s))
    return [convert(text, i % 2 == 1) for i, text in enumerate(parts)]


def run_clean_validate_with_clean_log(*args, **kwargs):

    buffer = io.StringIO()

    with redirect_stdout(buffer):
        result = cleanValidate(*args, **kwargs)

    printed_text = buffer.getvalue().strip()

    for line in printed_text.splitlines():
        if line.startswith("Data Clean Report file:"):
            report_path = line.replace("Data Clean Report file:", "").strip()
            logging.info(f"Data Clean Report file created: {os.path.basename(report_path)}")
        elif line.strip():
            logging.info(line.strip())

    return result
DATE_ERROR_LIMIT = 3


def _date_field_display_name(field_name):
    text = str(field_name or "").strip()
    idx = 0

    while idx < len(text) and (text[idx].isdigit() or text[idx] == "."):
        idx += 1

    return text[idx:] or text


def _job_files(project_path, original_filename, fmt_name):
    base_name, _ = os.path.splitext(original_filename)
    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")
    report_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Report.xlsx")
    working_file = os.path.join(project_path, f"{base_name}_working.xlsx")
    date_error_file = os.path.join(project_path, "date_errors.json")

    return base_name, cleaned_file, report_file, working_file, date_error_file


def _create_working_file(source_file, working_file):
    if os.path.abspath(source_file) != os.path.abspath(working_file):
        shutil.copy2(source_file, working_file)


def _build_date_errors(sorted_df, sorted_mask, alias_mapping, date_error_file):
    errors = []

    for row_pos, (source_row_index, mask_row) in enumerate(sorted_mask.iterrows()):
        date_cols = [
            col for col, val in mask_row.items()
            if val == "dateformat"
        ]

        if len(date_cols) < DATE_ERROR_LIMIT:
            continue

        row_data = sorted_df.iloc[row_pos]
        _, msgs = validate_date_rules(row_data, alias_mapping)
        field_names = [_date_field_display_name(col) for col in date_cols]
        display_msgs = [
            msg for msg in msgs
            if any(field_name and field_name in msg for field_name in field_names)
        ]

        errors.append({
            "row_index": row_pos,
            "source_row_index": int(source_row_index),
            "excel_row": int(source_row_index) + 2,
            "clean_excel_row": row_pos + 2,
            "fields": [
                {
                    "name": col,
                    "value": "" if pd.isna(row_data.get(col, "")) else str(row_data.get(col, ""))
                }
                for col in date_cols
            ],
            "messages": display_msgs or msgs or ["日期邏輯錯誤"]
        })

    with open(date_error_file, "w", encoding="utf-8") as f:
        json.dump(errors, f, ensure_ascii=False, indent=2)

    return errors


def _resolve_source_row_index(date_error_file, row_index):
    if not os.path.exists(date_error_file):
        return row_index

    try:
        with open(date_error_file, "r", encoding="utf-8") as f:
            date_errors = json.load(f)
    except (json.JSONDecodeError, OSError):
        return row_index

    for item in date_errors:
        if int(item.get("row_index", -1)) == int(row_index):
            return int(item.get("source_row_index", row_index))

    return row_index


def _load_job(job_id):
    conn = get_conn()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            Job.Path,
            Job.FileName,
            DataFormat.FmtName,
            DataFormat.Version,
            DataFormat.Revision_date
        FROM [Job]
        JOIN [DataFormat]
            ON Job.FmtID = DataFormat.FmtID
        WHERE Job.JobID = ?
    """, (job_id,))

    row = cursor.fetchone()
    conn.close()

    return row


def _update_working_file_cell_only(working_file, row_index, updates):
    wb = load_workbook(working_file)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_idx

    excel_row = int(row_index) + 2

    for field_name, new_value in updates.items():
        if field_name not in headers:
            continue

        col_idx = headers[field_name]
        cell = ws.cell(row=excel_row, column=col_idx)

        cell.value = (
            "" if new_value is None
            else str(new_value).replace("-", "").replace("/", "").strip()
        )

        # 只把被修改的日期欄位設成文字格式
        cell.number_format = "@"

    wb.save(working_file)
    wb.close()


def _detect_text_file_format(file_path):
    encodings = ["utf-8-sig", "utf-8", "cp950", "big5"]
    content = None
    used_encoding = "utf-8-sig"

    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                content = f.read(4096)
            used_encoding = enc
            break
        except UnicodeDecodeError:
            continue

    if content is None:
        content = ""

    first_line = content.splitlines()[0] if content.splitlines() else ""

    if "\t" in first_line:
        delimiter = "\t"
    elif "," in first_line:
        delimiter = ","
    elif ";" in first_line:
        delimiter = ";"
    else:
        ext = os.path.splitext(file_path)[1].lower()
        delimiter = "," if ext == ".csv" else "\t"

    return used_encoding, delimiter


def _cell_text(value):
    if value is None:
        return ""

    text = str(value).strip()

    if text.endswith(".0"):
        text = text[:-2]

    return text


def _working_file_rows(working_file):
    wb = load_workbook(working_file, data_only=False)
    ws = wb.active

    rows = [
        [_cell_text(value) for value in row]
        for row in ws.iter_rows(values_only=True)
    ]

    wb.close()
    return rows


def _truncate_to_encoded_width(text, width, encoding):
    result = ""

    for ch in text:
        candidate = result + ch
        if len(candidate.encode(encoding, errors="ignore")) > width:
            break
        result = candidate

    return result


def _format_fixed_width_value(value, width, encoding):
    text = _truncate_to_encoded_width(_cell_text(value), width, encoding)
    byte_len = len(text.encode(encoding, errors="ignore"))
    return text + (" " * max(width - byte_len, 0))


def _write_fixed_width_txt_from_working(working_file, source_file, fmt_name, encoding):
    fmt_val = str(fmt_name).replace("fmt_", "")
    field_spec = load_field_spec(fmt_val)
    rows = _working_file_rows(working_file)

    if not rows:
        return

    headers = rows[0]
    header_index = {name: idx for idx, name in enumerate(headers)}
    output_lines = []

    for row in rows[1:]:
        line_parts = []

        for field_name, start, end in field_spec:
            col_idx = header_index.get(field_name)
            value = row[col_idx] if col_idx is not None and col_idx < len(row) else ""
            line_parts.append(
                _format_fixed_width_value(value, end - start + 1, encoding)
            )

        output_lines.append("".join(line_parts))

    with open(source_file, "w", encoding=encoding, newline="") as f:
        f.write("\n".join(output_lines))
        if output_lines:
            f.write("\n")


def _write_csv_from_working(working_file, source_file, encoding, delimiter):
    rows = _working_file_rows(working_file)

    with open(source_file, "w", encoding=encoding, newline="") as f:
        writer = csv.writer(
            f,
            delimiter=delimiter,
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writerows(rows)


def _sync_working_file_to_source_files(working_file, source_file, converted_file, fmt_name):
    if not os.path.exists(working_file):
        return

    working_abs = os.path.abspath(working_file)

    if converted_file and os.path.exists(converted_file):
        converted_abs = os.path.abspath(converted_file)
        if converted_abs != working_abs:
            shutil.copy2(working_file, converted_file)

    if not source_file or not os.path.exists(source_file):
        return

    source_abs = os.path.abspath(source_file)
    ext = os.path.splitext(source_file)[1].lower()

    if ext in [".xlsx", ".xlsm"] and source_abs != working_abs:
        shutil.copy2(working_file, source_file)
        return

    if ext not in [".txt", ".csv"]:
        return

    encoding, delimiter = _detect_text_file_format(source_file)

    if ext == ".csv":
        _write_csv_from_working(working_file, source_file, encoding, delimiter)
        return

    _write_fixed_width_txt_from_working(
        working_file,
        source_file,
        fmt_name,
        encoding,
    )


def _build_cleaning_analysis(stats, sorted_mask):
    by_field = []

    for col in sorted_mask.columns:
        err_count = (sorted_mask[col] != "").sum()

        if err_count > 0:
            by_field.append({
                "name": col,
                "format": "檢核不符",
                "errors": int(err_count)
            })

    total_cells = stats["total"] * len(sorted_mask.columns)

    by_type = [
        {
            "type": "A:遺漏值",
            "count": int(stats["missing_cells"]),
            "ratio": f"{(stats['missing_cells'] / total_cells * 100):.1f}%" if total_cells > 0 else "0%"
        },
        {
            "type": "B:格式不符",
            "count": int(stats["format_cells"]),
            "ratio": f"{(stats['format_cells'] / total_cells * 100):.1f}%" if total_cells > 0 else "0%"
        },
        {
            "type": "C:邏輯錯誤",
            "count": int(stats["logic_cells"]),
            "ratio": f"{(stats['logic_cells'] / total_cells * 100):.1f}%" if total_cells > 0 else "0%"
        }
    ]

    return {
        "by_field": by_field,
        "by_type": by_type
    }


def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def load_field_spec(fmt_val):
    conn = get_conn()
    cursor = conn.cursor()
    
    sql_fields = """SELECT [ChineseName], [Start], [End] FROM [CancerRegistry_Fields] 
                   WHERE [fmt]=? ORDER BY [Start]"""
    cursor.execute(sql_fields, fmt_val)
    field_rows = cursor.fetchall()
    
    sql_map = "SELECT [序號], [中文欄位名稱] FROM [CancerRegistry_FieldMap]"
    cursor.execute(sql_map)
    map_rows = cursor.fetchall()
    conn.close()
    
    name_to_seq = {}
    for r_seq, r_name in map_rows:
        if not r_seq or not r_name: continue
        seq_str = str(r_seq).strip().replace('.0', '')
        full_name = str(r_name).strip()
        
        name_to_seq[full_name] = seq_str
        
        if seq_str in ['4.2.1.8', '7.6'] and '/' in full_name:
            for part in full_name.split('/'):
                part = part.strip()
                if part:
                    name_to_seq[part] = seq_str

    from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
    fmt_key = f"fmt_{fmt_val}"
    rules = FORMAT_RULES_MAP.get(fmt_key, {})
    id_to_rule_name = {r_val.get('ID'): r_key for r_key, r_val in rules.items() if r_val.get('ID')}
    
    field_spec = []
    for f_name, f_start, f_end in field_rows:
        orig_name = str(f_name).strip()
        
        seq = name_to_seq.get(orig_name, "")
        
        target_name = orig_name
        if seq in ['4.2.1.8', '7.6']:
            rule_name = id_to_rule_name.get(seq)
            if rule_name:
                for m_seq, m_full in map_rows:
                    if str(m_seq).strip().replace('.0', '') == seq:
                        if rule_name in str(m_full):
                            target_name = rule_name
                            break

        header_name = f"{seq}{target_name}" if seq else target_name
        field_spec.append((header_name, int(f_start), int(f_end)))
        
    return field_spec

def parse_fixed_width_line(line_text, spec):
    line_bytes = line_text.encode('big5', errors='ignore')
    parsed_row = {}
    for name, start, end in spec:
        field_value = line_bytes[start - 1:end]
        try:
            parsed_row[name] = field_value.decode('big5').strip()
        except:
            parsed_row[name] = field_value.decode('big5', errors='replace').strip()
    return parsed_row


def categorize_fields_logic(job_id, scheme):
    if not job_id or not scheme:
        return {"ok": False, "error": "參數不足"}, 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID WHERE Job.JobID=?", (job_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return {"ok": False, "error": "找不到該紀錄"}, 404

    project_path, original_filename, fmt_name = row
    base_name, _ = os.path.splitext(original_filename)
    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")

    if not os.path.exists(cleaned_file):
        return {"ok": False, "error": "清洗檔案不存在"}, 404

    try:
        wb = load_workbook(cleaned_file, read_only=True)
        ws = wb.active
        rows_gen = ws.iter_rows(max_row=1)
        first_row = next(rows_gen, None)
        if first_row is None:
            headers = []
        else:
            headers = [str(cell.value).strip() if cell.value else "" for cell in first_row]
        wb.close()

        from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
        fmt_key = f"fmt_{fmt_name}" if not str(fmt_name).startswith("fmt_") else str(fmt_name)
        rules = FORMAT_RULES_MAP.get(fmt_key, {})
        id_to_rule_name = {v.get("ID"): k for k, v in rules.items() if v.get("ID")}

        alias_to_target = get_field_map(scheme, fmt_name)
        mapped = []
        unmapped = []

        if scheme == 'original':
            import json
            orig_headers_path = os.path.join(project_path, "original_headers.json")
            if os.path.exists(orig_headers_path):
                with open(orig_headers_path, "r", encoding="utf-8") as f_orig:
                    orig_headers = json.load(f_orig)
            else:
                orig_headers = headers

            alias_to_target_zh = get_field_map("field_name_zh", fmt_name)
            for idx, h in enumerate(headers):
                if not h or h.startswith('_') or h == '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)':
                    continue
                orig_label = orig_headers[idx] if idx < len(orig_headers) else h
                if h in alias_to_target_zh:
                    mapped.append({"key": h, "label": orig_label, "target": orig_label})
                else:
                    unmapped.append({"key": h, "label": orig_label})
        else:
            from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
            fmt_key = f"fmt_{fmt_name}" if not str(fmt_name).startswith("fmt_") else str(fmt_name)
            rules = FORMAT_RULES_MAP.get(fmt_key, {})
            id_to_rule_name = {v.get("ID"): k for k, v in rules.items() if v.get("ID")}

            alias_to_target = get_field_map(scheme, fmt_name)
            for h in headers:
                if not h or h.startswith('_') or h == '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)':
                    continue
                if h in alias_to_target:
                    target_name = alias_to_target[h]
                    final_display = target_name
                    
                    if (scheme == 'field_name_zh' or scheme == '中文欄位名稱') and '/' in target_name:
                        m_target = re.match(r'^(\d+(\.\d+)*)', target_name)
                        if m_target:
                            seq = m_target.group(1)
                            if seq in ['4.2.1.8', '7.6']:
                                target_raw = target_name[len(seq):].strip()
                                valid_parts = [p.strip() for p in target_raw.split('/') if p.strip()]
                                source_raw = h[len(seq):].strip() if h.startswith(seq) else h
                                
                                if source_raw in valid_parts:
                                    final_display = f"{seq}{source_raw}"
                                else:
                                    rule_name = id_to_rule_name.get(seq)
                                    if rule_name:
                                        final_display = f"{seq}{rule_name}"
                    
                    mapped.append({"key": h, "label": final_display, "target": final_display})
                else:
                    unmapped.append({"key": h, "label": h})
        
        mapped.sort(key=lambda x: _natural_sort_key(x['label']))
        unmapped.sort(key=lambda x: _natural_sort_key(x['label']))

        return {"ok": True,"mapped": mapped,"unmapped": unmapped}, 200
    except Exception as e:
        return {"ok": False, "error": str(e)}, 500

def export_logic(job_id, scheme, selected_fields):
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID WHERE Job.JobID=?", (job_id,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return {"ok": False, "error": "找不到該紀錄"}, 404

    project_path, original_filename, fmt_name = row
    base_name, _ = os.path.splitext(original_filename)
    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")

    if not os.path.exists(cleaned_file):
        return {"ok": False, "error": "清洗檔案不存在"}, 404

    alias_to_target = get_field_map(scheme, fmt_name)
    
    from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
    fmt_key = f"fmt_{fmt_name}" if not str(fmt_name).startswith("fmt_") else str(fmt_name)
    rules = FORMAT_RULES_MAP.get(fmt_key, {})
    id_to_rule_name = {v.get("ID"): k for k, v in rules.items() if v.get("ID")}

    scheme_display = {
        "original":"原始匯入欄位名稱",
        "field_name_zh":"中文欄位名稱",
        "field_name_en":"英文欄位名稱",
        "ntu_yunlin":"台大雲林欄位名稱",
        "ntu_system":"台大體系醫整庫欄位名稱",
        "taiwan_cancer_registry":"台灣癌症登記中心",
        "AI_module":"雲醫癌AI模組"}.get(scheme, scheme)

    try:
        wb = load_workbook(cleaned_file)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        output_cols = []
        used_orig_indices = set()

        if scheme == 'original':
            import json
            orig_headers_path = os.path.join(project_path, "original_headers.json")
            if os.path.exists(orig_headers_path):
                with open(orig_headers_path, "r", encoding="utf-8") as f_orig:
                    orig_headers = json.load(f_orig)
            else:
                orig_headers = headers

            alias_to_target_zh = get_field_map("field_name_zh", fmt_name)
            unique_targets = sorted(list(set(alias_to_target_zh.values())), key=_natural_sort_key)

            for target_name in unique_targets:
                source_idx = None
                for i, h in enumerate(headers):
                    if h == target_name or (h in alias_to_target_zh and alias_to_target_zh[h] == target_name):
                        source_idx = i + 1
                        break
                if source_idx:
                    orig_label = orig_headers[source_idx - 1] if (source_idx - 1) < len(orig_headers) else target_name
                    output_cols.append((source_idx, orig_label))
                    used_orig_indices.add(source_idx)

            for field in selected_fields:
                if field in headers:
                    orig_idx = headers.index(field) + 1
                    if orig_idx not in used_orig_indices:
                        orig_label = orig_headers[orig_idx - 1] if (orig_idx - 1) < len(orig_headers) else field
                        output_cols.append((orig_idx, orig_label))
                        used_orig_indices.add(orig_idx)
        else:
            alias_to_target = get_field_map(scheme, fmt_name)
            
            from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
            fmt_key = f"fmt_{fmt_name}" if not str(fmt_name).startswith("fmt_") else str(fmt_name)
            rules = FORMAT_RULES_MAP.get(fmt_key, {})
            id_to_rule_name = {v.get("ID"): k for k, v in rules.items() if v.get("ID")}

            unique_targets = sorted(list(set(alias_to_target.values())), key=_natural_sort_key)

            for target_name in unique_targets:
                source_idx = None
                final_display_name = target_name
                
                for i, h in enumerate(headers):
                    if h in alias_to_target and alias_to_target[h] == target_name:
                        source_idx = i + 1
                        
                        if scheme_display == '中文欄位名稱' and '/' in target_name:
                            m_target = re.match(r'^(\d+(\.\d+)*)', target_name)
                            if m_target:
                                seq_prefix = m_target.group(1)
                                if seq_prefix in ['4.2.1.8', '7.6']:
                                    target_raw_name = target_name[len(seq_prefix):].strip()
                                    valid_parts = [p.strip() for p in target_raw_name.split('/') if p.strip()]
                                    
                                    rule_name = id_to_rule_name.get(seq_prefix)
                                    if rule_name:
                                        final_display_name = f"{seq_prefix}{rule_name}"

                                    if h.startswith(seq_prefix):
                                        source_raw_name = h[len(seq_prefix):].strip()
                                        if source_raw_name in valid_parts:
                                            final_display_name = h
                        break

                if source_idx:
                    output_cols.append((source_idx, final_display_name))
                    used_orig_indices.add(source_idx)

            for field in selected_fields:
                if field in headers:
                    orig_idx = headers.index(field) + 1
                    if orig_idx not in used_orig_indices:
                        output_cols.append((orig_idx, field))
                        used_orig_indices.add(orig_idx)

        err_col_name = '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)'
        if err_col_name in headers:
            err_idx = headers.index(err_col_name) + 1
            if err_idx not in used_orig_indices:
                output_cols.append((err_idx, err_col_name))


        new_wb = Workbook()
        new_ws = new_wb.active
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            for c_idx, (orig_col_idx, display_name) in enumerate(output_cols, start=1):
                target_cell = new_ws.cell(row=r_idx, column=c_idx)
                
                if orig_col_idx is not None:
                    source_cell = ws.cell(row=r_idx, column=orig_col_idx)
                    if r_idx == 1:
                        target_cell.value = display_name
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.alignment = copy(source_cell.alignment)
                    else:
                        copy_cell(source_cell, target_cell)
                else:
                    if r_idx == 1:
                        target_cell.value = display_name
                    else:
                        target_cell.value = ""

        for c_idx, (orig_col_idx, _) in enumerate(output_cols, start=1):
            if orig_col_idx is not None:
                new_ws.column_dimensions[get_column_letter(c_idx)].width =                     ws.column_dimensions[get_column_letter(orig_col_idx)].width
            else:
                new_ws.column_dimensions[get_column_letter(c_idx)].width = 15

        export_filename = f"fmt{fmt_name}_{base_name}_{scheme_display}.xlsx"
        temp_out = os.path.join(project_path, export_filename)
        new_wb.save(temp_out)
        
        return {"send_file": True, "path": os.path.abspath(temp_out), "download_name": export_filename}, 200

    except Exception as e:
        return {"ok": False, "error": str(e)}, 500

def preview_logic(job_id, scheme, selected_fields):
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID WHERE Job.JobID=?", (job_id,))
    row = cursor.fetchone()
    conn.close()

    project_path, original_filename, fmt_name = row
    base_name, _ = os.path.splitext(original_filename)
    cleaned_file = os.path.join(project_path, f"fmt{fmt_name}_{base_name}_Clean.xlsx")

    if not os.path.exists(cleaned_file):
        return {"ok": False, "error": "清洗檔案不存在"}, 404
    alias_to_target = get_field_map(scheme, fmt_name)
    
    from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
    fmt_key = f"fmt_{fmt_name}" if not str(fmt_name).startswith("fmt_") else str(fmt_name)
    rules = FORMAT_RULES_MAP.get(fmt_key, {})
    id_to_rule_name = {v.get("ID"): k for k, v in rules.items() if v.get("ID")}

    try:
        wb = load_workbook(cleaned_file, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        output_cols = []
        used_orig_indices = set()

        if scheme == 'original':
            import json
            orig_headers_path = os.path.join(project_path, "original_headers.json")
            if os.path.exists(orig_headers_path):
                with open(orig_headers_path, "r", encoding="utf-8") as f_orig:
                    orig_headers = json.load(f_orig)
            else:
                orig_headers = headers

            alias_to_target_zh = get_field_map("field_name_zh", fmt_name)
            unique_targets = sorted(list(set(alias_to_target_zh.values())), key=_natural_sort_key)

            for target_name in unique_targets:
                source_idx = None
                for i, h in enumerate(headers):
                    if h == target_name or (h in alias_to_target_zh and alias_to_target_zh[h] == target_name):
                        source_idx = i + 1
                        break
                if source_idx:
                    orig_label = orig_headers[source_idx - 1] if (source_idx - 1) < len(orig_headers) else target_name
                    output_cols.append((source_idx, orig_label))
                    used_orig_indices.add(source_idx)

            for f in selected_fields:
                if f in headers:
                    idx = headers.index(f) + 1
                    if idx not in used_orig_indices:
                        orig_label = orig_headers[idx - 1] if (idx - 1) < len(orig_headers) else f
                        output_cols.append((idx, orig_label))
                        used_orig_indices.add(idx)
        else:
            alias_to_target = get_field_map(scheme, fmt_name)
            
            from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
            fmt_key = f"fmt_{fmt_name}" if not str(fmt_name).startswith("fmt_") else str(fmt_name)
            rules = FORMAT_RULES_MAP.get(fmt_key, {})
            id_to_rule_name = {v.get("ID"): k for k, v in rules.items() if v.get("ID")}

            unique_targets = sorted(list(set(alias_to_target.values())), key=_natural_sort_key)

            for target_name in unique_targets:
                source_idx = None
                final_display_name = target_name
                
                for i, h in enumerate(headers):
                    if h in alias_to_target and alias_to_target[h] == target_name:
                        source_idx = i + 1
                        
                        if (scheme == 'field_name_zh' or scheme == '中文欄位名稱') and '/' in target_name:
                            m_target = re.match(r'^(\d+(\.\d+)*)', target_name)
                            if m_target:
                                seq_prefix = m_target.group(1)
                                if seq_prefix in ['4.2.1.8', '7.6']:
                                    target_raw_name = target_name[len(seq_prefix):].strip()
                                    valid_parts = [p.strip() for p in target_raw_name.split('/') if p.strip()]
                                    
                                    rule_name = id_to_rule_name.get(seq_prefix)
                                    if rule_name:
                                        final_display_name = f"{seq_prefix}{rule_name}"

                                    if h.startswith(seq_prefix):
                                        source_raw_name = h[len(seq_prefix):].strip()
                                        if source_raw_name in valid_parts:
                                            final_display_name = h
                            break
                
                if source_idx:
                    output_cols.append((source_idx, final_display_name))
                    used_orig_indices.add(source_idx)

            for f in selected_fields:
                if f in headers:
                    idx = headers.index(f) + 1
                    if idx not in used_orig_indices:
                        output_cols.append((idx, f))
                        used_orig_indices.add(idx)

        err_f = '錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)'
        if err_f in headers:
            idx = headers.index(err_f) + 1
            if idx not in used_orig_indices:
                output_cols.append((idx, err_f))

        preview_data = []
        preview_headers = [col[1] for col in output_cols]
        for r_idx in range(2, 7):
            row_data = []
            for c_idx, _ in output_cols:
                if c_idx is not None:
                    val = ws.cell(row=r_idx, column=c_idx).value
                    row_data.append(str(val or ""))
                else:
                    row_data.append("")
            preview_data.append(row_data)
        return {"ok": True,"headers": preview_headers,"data": preview_data}, 200

    except Exception as e:
        return {"ok": False, "error": str(e)}, 500

def get_formats_logic():
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT [FmtID], [FmtName], [Version], [Revision_date] FROM [DataFormat] ORDER BY FmtName ASC")
    formats = [{"id": str(r.FmtID), "name": str(r.FmtName), "version": str(r.Version), "updated": str(r.Revision_date)} for r in cursor.fetchall()]
    conn.close()
    return formats

def add_format_logic(name, version, updated):
    if not name or not version: return {"ok": False, "message": "名稱與版本為必填"}, 400
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO [DataFormat] ([FmtName], [Version], [Revision_date]) VALUES (?, ?, ?)", (name, version, updated))
    conn.commit()
    conn.close()
    return {"ok": True}, 200

def manage_format_logic(method, fmt_id, name, version, updated):
    conn = get_conn()
    cursor = conn.cursor()

    if method == "DELETE":
        cursor.execute("DELETE FROM [DataFormat] WHERE [FmtID] = ?", (fmt_id,))
        conn.commit()
        conn.close()
        return {"ok": True}, 200

    cursor.execute("UPDATE [DataFormat] SET [FmtName]=?, [Version]=?, [Revision_date]=? WHERE [FmtID]=?", (name, version, updated, fmt_id))
    conn.commit()
    conn.close()
    return {"ok": True}, 200

def clean_job_logic(user_id, format_id, convert_txt_flag, uploaded_file):
    if not format_id or not uploaded_file or uploaded_file.filename == '': 
        return {"ok": False, "error": "未選擇檔案"}, 400
    
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT [FmtName], [Version], [Revision_date] FROM [DataFormat] WHERE [FmtID] = ?", (format_id,))
    fmt_data = cursor.fetchone()
    if not fmt_data:
        conn.close()
        return {"ok": False, "error": "找不到指定的格式"}, 400
        
    fmt_name, version, rev_date = fmt_data
    filename = os.path.basename(uploaded_file.filename)
    if not filename: filename = "uploaded_file"
    file_ext = os.path.splitext(filename)[1].lower()
    base_name = os.path.splitext(filename)[0]
    has_no_headers = (file_ext == '.txt' and not convert_txt_flag)

    if file_ext == '.txt':
        uploaded_file.seek(0)
        content_bytes = uploaded_file.read()
        uploaded_file.seek(0)

        if convert_txt_flag:
            try:
                text_sample = content_bytes[:8192].decode('utf-8', errors='ignore')
                first_line = text_sample.splitlines()[0] if text_sample else ""
                delimiter = "\t" if "\t" in first_line else ("," if "," in first_line else ";")
                
                found_headers = [h.strip() for h in first_line.split(delimiter)]
                
                from modules.blueprint.clean.cleaner import FORMAT_RULES_MAP
                norm_fmt = f"fmt_{str(fmt_name).replace('fmt_', '')}"
                rules = FORMAT_RULES_MAP.get(norm_fmt, {})
                valid_std_headers = set()
                for rule_name, rule_val in rules.items():
                    r_id = str(rule_val.get("ID", "")).strip()
                    if r_id:
                        valid_std_headers.add(f"{r_id}{rule_name}")

                has_any_match = False
                for h in found_headers:
                    try:
                        renamed = validate_and_rename_headers([h], fmt_name)
                        if renamed and renamed[0] in valid_std_headers:
                            has_any_match = True
                            break
                    except ValueError:
                        continue
                
                if not has_any_match:
                    conn.close()
                    return {
                        "ok": False, 
                        "error": "標頭偵測失敗",
                        "message": "檔案可能不包含標頭列，請確認是否誤勾選「包含首行標頭」，或檔案內容格式不符。"
                    }, 400
            except Exception as e:
                conn.close()
                return {"ok": False, "error": f"標頭驗證失敗: {str(e)}"}, 500
        else:
            fmt_val = str(fmt_name).replace("fmt_", "")
            cursor.execute("SELECT MAX([End]) FROM CancerRegistry_Fields WHERE [fmt]=? GROUP BY [fmt]", (fmt_val,))
            row = cursor.fetchone()
            expected_length = row[0] if row else 0

            length_errors = []
            lines = content_bytes.splitlines()
            for i, line_bytes in enumerate(lines):
                if not line_bytes.strip(): continue
                length = len(line_bytes)
                if expected_length > 0 and length != expected_length:
                    length_errors.append(f"第 {i+1} 行: 實際 {length}, 預期 {expected_length}")
            if length_errors:
                conn.close()
                
                log_content = f"檔案名稱: {filename}\n" + "\n".join(length_errors)
                log_base64 = base64.b64encode(log_content.encode('utf-8')).decode('utf-8')
                
                field_spec = load_field_spec(fmt_val)
                results = []
                for line_bytes in lines:
                    results.append(parse_fixed_width_line(line_bytes.decode('big5', errors='ignore'), field_spec))
                
                output_xlsx = io.BytesIO()
                keys = [f[0] for f in field_spec]
                wb = Workbook()
                ws = wb.active
                ws.append(keys)
                for r in results: ws.append([r.get(k, "") for k in keys])
                for row in ws.iter_rows():
                    for cell in row: cell.number_format = '@'
                wb.save(output_xlsx)
                xlsx_base64 = base64.b64encode(output_xlsx.getvalue()).decode('utf-8')

                error_message_html = f"檔案長度校驗失敗，總共有 <b>{len(length_errors)}</b> 筆長度不符的資料：<br>"
                error_message_html += "<br>".join(length_errors[:3])
                if len(length_errors) > 3: error_message_html += f"<br>...以及其他 {len(length_errors) - 3} 筆錯誤"
                
                return {
                    "ok": False, 
                    "error": error_message_html, 
                    "has_length_error": True,
                    "log_data": log_base64,
                    "xlsx_data": xlsx_base64,
                    "filename": filename
                }, 400

    JobID = str(uuid.uuid4())
    Jobs_FOLDER = 'tasks/Jobs'
    project_folder = f"{Jobs_FOLDER}/{JobID}"
    os.makedirs(project_folder, exist_ok=True)

    path = f"{project_folder}/{filename}"
    uploaded_file.save(path)
    uploaded_file.close()
    process_path = path

    if file_ext == '.csv':
        try:
            df_csv = pd.read_csv(path, dtype=str, encoding='utf-8-sig')
        except UnicodeDecodeError:
            df_csv = pd.read_csv(path, dtype=str, encoding='cp950')
        temp_xlsx = f"{project_folder}/{base_name}.xlsx"
        df_csv.to_excel(temp_xlsx, index=False)
        process_path = temp_xlsx

    if file_ext == '.txt':
        if convert_txt_flag:
            try:
                temp_xlsx = f"{project_folder}/{base_name}.xlsx"
                convert_txt_to_excel(path, temp_xlsx)
                if os.path.exists(temp_xlsx):
                    process_path = temp_xlsx
                else: raise Exception("Excel conversion failed")
            except Exception as e:
                if conn and not conn.closed: conn.close()
                return {"ok": False, "error": f"TXT 轉換失敗: {str(e)}"}, 500
        else:
            try:
                fmt_val = str(fmt_name).replace("fmt_", "")
                field_spec = load_field_spec(fmt_val)
                results = []
                with open(path, 'r', encoding='big5', errors='ignore') as f:
                    for line in f: results.append(parse_fixed_width_line(line, field_spec))
                temp_xlsx = f"{project_folder}/{base_name}.xlsx"
                keys = [f[0] for f in field_spec]
                wb = Workbook()
                ws = wb.active
                ws.append(keys)
                for r in results: ws.append([r.get(k, "") for k in keys])
                for row in ws.iter_rows():
                    for cell in row: cell.number_format = '@'
                wb.save(temp_xlsx)
                process_path = temp_xlsx
            except Exception as e:
                if conn and not conn.closed: conn.close()
                return {"ok": False, "error": f"TXT 解析失敗: {str(e)}"}, 500

    base_name, out_path, rep_path, working_file, date_error_file = _job_files(project_folder, filename, fmt_name)
    
    try:
        wb_orig = load_workbook(process_path, read_only=True)
        orig_ws = wb_orig.active
        orig_rows_gen = orig_ws.iter_rows(max_row=1)
        orig_first_row = next(orig_rows_gen, None)
        if orig_first_row is None:
            orig_headers = []
        else:
            orig_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in orig_first_row]
        wb_orig.close()
        
        import json
        with open(os.path.join(project_folder, "original_headers.json"), "w", encoding="utf-8") as f_orig:
            json.dump(orig_headers, f_orig, ensure_ascii=False)

        validate_and_unify_headers_in_file(process_path, fmt_name)
        
        _create_working_file(process_path, working_file)
        stats, alias_mapping, sorted_df, sorted_mask = cleanValidate(process_path, out_path, rep_path, f"fmt_{fmt_name}", version, rev_date)
        date_errors = _build_date_errors(sorted_df, sorted_mask, alias_mapping, date_error_file)
        cursor.execute("INSERT INTO Job ([JobID],[UserID],[FmtID],[FileName],[TotalCount],[CompletenessScore],[CorrectScore],[ConsistencyScore],[DQI],[Path]) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (JobID, user_id, format_id, filename, int(stats['total']), float(stats['completeness']), float(stats['correctness']), float(stats['consistency']), float(stats['quality_score']), project_folder))
        conn.commit()
    except ValueError as e:
        if conn and not conn.closed: conn.close()
        return {"ok": False, "error": str(e)}, 400
    except Exception as e:
        if conn and not conn.closed: conn.rollback()
        return {"ok": False, "error":str(e)}, 500
    finally: 
        if conn and not conn.closed:
            conn.close()

    by_field = []
    for col in sorted_mask.columns:
        err_count = (sorted_mask[col] != "").sum()
        if err_count > 0:
            by_field.append({"name": col,"format": "檢核不符","errors": int(err_count)})
    
    by_type = [
        {"type": "A:遺漏值", "count": int(stats['missing_cells']), "ratio": f"{(stats['missing_cells']/(stats['total']*len(sorted_mask.columns))*100):.1f}%" if stats['total'] > 0 else "0%"},
        {"type": "B:格式不符", "count": int(stats['format_cells']), "ratio": f"{(stats['format_cells']/(stats['total']*len(sorted_mask.columns))*100):.1f}%" if stats['total'] > 0 else "0%"},
        {"type": "C:邏輯錯誤", "count": int(stats['logic_cells']), "ratio": f"{(stats['logic_cells']/(stats['total']*len(sorted_mask.columns))*100):.1f}%" if stats['total'] > 0 else "0%"}
    ]

    output_fields = [{"key": col, "label": col} for col in sorted_df.columns if not col.startswith('_')]
    
    detected_system, _ = detect_system(orig_headers)

    date_error_count = len(date_errors)
    DATE_ERROR_LIMIT = 3

    if date_error_count > 0:
        message = (
            f"日期邏輯錯誤共有 {date_error_count} 筆，"
            f"已達系統限制 {DATE_ERROR_LIMIT} 筆，"
            "請先修正錯誤的日期資料，完成修正後再進行後續資料清洗作業"
        )
        can_continue = False
    else:
        message = "清洗完成"
        can_continue = True

    return {
        "ok": True,
        "project_id": JobID,
        "job_id": JobID,
        "detected_system": detected_system,
        "has_no_headers": has_no_headers,
        "message": message,
        "can_continue": can_continue,
        "stats": {
            "total": int(stats['total']), 
            "passed": int(stats['total'] - stats['error_rows']), 
            "error": int(stats['error_rows']), 
            "completeness": float(stats['completeness']), 
            "correctness": float(stats['correctness']), 
            "consistency": float(stats['consistency']),
            "dqi": float(stats['quality_score'])
        },
        "analysis": {"by_field": by_field,"by_type": by_type},
        "output_fields": output_fields,
        "date_error_limit": DATE_ERROR_LIMIT,
        "date_error_count": date_error_count,
        "date_errors": date_errors
    }, 200


def get_date_errors_logic(job_id):
    if not job_id:
        return {"ok": False, "error": "缺少 job_id"}, 400

    row = _load_job(job_id)

    if not row:
        return {"ok": False, "error": "找不到該紀錄"}, 404

    project_path, original_filename, fmt_name, version, rev_date = row

    _, _, _, _, date_error_file = _job_files(project_path, original_filename, fmt_name)

    DATE_ERROR_LIMIT = 3
    if not os.path.exists(date_error_file):
        return {
            "ok": True,
            "date_errors": [],
            "date_error_count": 0,
            "date_error_limit": DATE_ERROR_LIMIT
        }, 200

    with open(date_error_file, "r", encoding="utf-8") as f:
        date_errors = json.load(f)

    return {
        "ok": True,
        "date_errors": date_errors,
        "date_error_count": len(date_errors),
        "date_error_limit": DATE_ERROR_LIMIT
    }, 200

def update_date_error_logic(job_id, row_index, updates):
    if not job_id:
        return {"ok": False, "error": "缺少 job_id"}, 400

    if row_index is None or not isinstance(updates, dict) or not updates:
        return {"ok": False, "error": "缺少修正資料"}, 400

    row = _load_job(job_id)
    if not row:
        return {"ok": False, "error": "找不到該紀錄"}, 404

    project_path, original_filename, fmt_name, version, rev_date = row
    _, cleaned_file, report_file, working_file, date_error_file = _job_files(project_path,original_filename,fmt_name)

    if not os.path.exists(working_file):
        return {"ok": False, "error": "找不到可修正的暫存資料"}, 404

    try:
        row_index = int(row_index)
    except ValueError:
        return {"ok": False, "error": "修正列格式錯誤"}, 400

    DATE_ERROR_LIMIT = 3
    try:
        source_row_index = _resolve_source_row_index(date_error_file, row_index)
        _update_working_file_cell_only(working_file, source_row_index, updates)
        source_file = os.path.join(project_path, original_filename)
        base_name, _ = os.path.splitext(original_filename)
        converted_file = os.path.join(project_path, f"{base_name}.xlsx")

        _sync_working_file_to_source_files(working_file,source_file,converted_file,fmt_name)
        stats, alias_mapping, sorted_df, sorted_mask = run_clean_validate_with_clean_log(
            working_file,cleaned_file,report_file,f"fmt_{fmt_name}",version,rev_date)
        date_errors = _build_date_errors(sorted_df, sorted_mask, alias_mapping, date_error_file)
        date_error_count = len(date_errors)

        conn = get_conn()
        cursor = conn.cursor()
        cursor.execute("UPDATE [Job] SET TotalCount = ?,CompletenessScore = ?,CorrectScore = ?,ConsistencyScore = ?,DQI = ?WHERE JobID = ?", (
            int(stats["total"]),
            float(stats["completeness"]),
            float(stats["correctness"]),
            float(stats["consistency"]),
            float(stats["quality_score"]),
            job_id
        ))
        conn.commit()
        conn.close()

    except Exception as e:
        return {"ok": False, "error": str(e)}, 500

    return {
        "ok": True,
        "message": "修正完成，已重新檢核",
        "job_id": job_id,
        "project_id": job_id,
        "date_error_limit": DATE_ERROR_LIMIT,
        "date_error_count": date_error_count,
        "date_errors": date_errors,
        "analysis": _build_cleaning_analysis(stats, sorted_mask),
        "stats": {
            "total": int(stats["total"]),
            "passed": int(stats["total"] - stats["error_rows"]),
            "error": int(stats["error_rows"]),
            "completeness": float(stats["completeness"]),
            "correctness": float(stats["correctness"]),
            "consistency": float(stats["consistency"]),
            "dqi": float(stats["quality_score"])
        }
    }, 200

def download_temp_file_logic(file_type, temp_id, filename):
    try:
        base_name = os.path.splitext(filename)[0]
        temp_folder = f"data/temp/{temp_id}"
        
        if file_type == "length_log":
            target_filename = "length_errors.log"
            display_name = f"長度錯誤_{base_name}.log"
        elif file_type == "xlsx":
            target_filename = f"{base_name}.xlsx"
            display_name = f"欄位檢核表_{base_name}.xlsx"
        else:
            return {"ok": False, "error": "無效的下載類型"}, 400

        file_path = os.path.join(temp_folder, target_filename)
        if not os.path.exists(file_path):
            return {"ok": False, "error": "檔案不存在或已過期"}, 404
            
        return {"send_file": True, "path": os.path.abspath(file_path), "download_name": display_name}, 200
    except Exception as e:
        return {"ok": False, "error": str(e)}, 500

def download_file_logic(file_type, job_id):
    try:
        conn = get_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT Job.Path, Job.FileName, DataFormat.FmtName FROM [Job] JOIN [DataFormat] ON Job.FmtID = DataFormat.FmtID WHERE Job.JobID=?", (job_id,))
        row = cursor.fetchone()
        if not row: return {"ok": False, "error": "找不到該紀錄"}, 404
        project_path, original_filename, fmt_name = row
        base_name, _ = os.path.splitext(original_filename)
        
        display_name = None
        if file_type == "cleaned":
            target_filename = f"fmt{fmt_name}_{base_name}_Clean.xlsx"
        elif file_type == "report":
            target_filename = f"fmt{fmt_name}_{base_name}_Report.xlsx"
        elif file_type == "original":
            target_filename = original_filename
        elif file_type == "parsed" or file_type == "preview":
            target_filename = f"{base_name}.xlsx"
            display_name = f"欄位檢核表_{base_name}.xlsx"
        elif file_type == "length_log":
            target_filename = "length_errors.log"
            display_name = f"欄位檢核表_{base_name}.log"
        else:
            return {"ok": False, "error": "無效的下載類型"}, 400

        file_path = os.path.join(project_path, target_filename)
        conn.close()
        if not os.path.exists(file_path): 
            return {"ok": False, "error": "檔案不存在"}, 404
            
        return {"send_file": True, "path": os.path.abspath(file_path), "download_name": display_name if display_name else target_filename}, 200
    except Exception as e:
        return {"ok": False, "error": str(e)}, 500

