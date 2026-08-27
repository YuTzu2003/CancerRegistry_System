import os
import re

import pandas as pd
from openpyxl import load_workbook


INPUT_SCHEMES = {
    "original", "field_name_zh", "field_name_en", "ntu_yunlin",
    "ntu_system", "taiwan_cancer_registry", "AI_module",
}
SCHEME_INDEX = {
    "field_name_zh": 1,
    "field_name_en": 2,
    "ntu_yunlin": 3,
    "ntu_system": 4,
    "taiwan_cancer_registry": 5,
    "AI_module": 6,
}
SEQUENCE_PATTERN = re.compile(r"^\d+(?:\.\d+)*$")
SEQUENCE_NAME_PATTERN = re.compile(r"^(\d+(?:\.\d+)*)\s*(.+)$")


def _text(value):
    return "" if value is None else str(value).strip()


def _normalize_sequence(value):
    text = _text(value)
    return text[:-2] if text.endswith(".0") else text


def _read_headers(file_path, extension):
    if extension == "xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return [_text(cell.value) for cell in workbook.active[1]]
        finally:
            workbook.close()
    try:
        return [_text(value) for value in pd.read_excel(file_path, nrows=0).columns]
    except ImportError as error:
        raise ValueError("舊式 .xls 檔案需要額外讀取元件，請先另存為 .xlsx 後再上傳。") from error


def _load_field_map(connection_factory):
    if connection_factory is None:
        raise ValueError("系統無法取得欄位名稱對照資料。")
    connection = connection_factory()
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT [序號], [中文欄位名稱], [英文欄位名稱], [台大雲林欄位名稱],
                   [台大體系醫整庫欄位名稱], [台灣癌症登記中心], [雲醫癌AI模組]
            FROM [Hospital_data].[dbo].[CancerRegistry_FieldMap]
        """)
        rows = []
        for row in cursor.fetchall():
            values = [_text(value) for value in row]
            values[0] = _normalize_sequence(values[0])
            if values[0] and values[1]:
                rows.append(values)
        return rows
    finally:
        connection.close()


def _compact(value):
    return re.sub(r"\s+", "", _text(value)).casefold()


def _maps_for_scheme(rows, input_scheme):
    by_sequence = {row[0]: f"{row[0]}{row[1]}" for row in rows}
    alias_map = {}
    indexes = range(1, 7) if input_scheme == "original" else (SCHEME_INDEX[input_scheme],)
    for row in rows:
        canonical = by_sequence[row[0]]
        for index in indexes:
            alias = _compact(row[index])
            if alias:
                alias_map.setdefault(alias, canonical)
    return by_sequence, alias_map


def _normalize_headers(headers, input_scheme, rows):
    empty_positions = [str(index + 1) for index, header in enumerate(headers) if not _text(header)]
    if empty_positions:
        raise ValueError(f"第一列表頭不可為空白（第 {', '.join(empty_positions)} 欄）。")

    by_sequence, alias_map = _maps_for_scheme(rows, input_scheme)
    normalized = []
    unknown_sequences = []
    for header in headers:
        text = _text(header)
        sequence_match = SEQUENCE_NAME_PATTERN.fullmatch(text)

        # 三種輸入形式都會自動辨識：僅序號、序號＋名稱、僅名稱。
        if SEQUENCE_PATTERN.fullmatch(text):
            sequence = _normalize_sequence(text)
            canonical = by_sequence.get(sequence)
            if not canonical:
                unknown_sequences.append(text)
                canonical = text
        elif sequence_match:
            sequence = _normalize_sequence(sequence_match.group(1))
            canonical = by_sequence.get(sequence)
            if not canonical:
                unknown_sequences.append(sequence)
                canonical = text
        else:
            # 未匹配的純名稱視為額外欄位並原樣保留。
            canonical = alias_map.get(_compact(text), text)
        normalized.append(canonical)

    if unknown_sequences:
        shown = ", ".join(dict.fromkeys(unknown_sequences[:5]))
        raise ValueError(f"找不到以下欄位序號的名稱對照：{shown}")
    if len(set(normalized)) != len(normalized):
        raise ValueError("欄位名稱轉換後出現重複欄位，請確認第一列表頭與命名來源。")
    return normalized


def _write_headers(file_path, extension, headers, keep_indexes=None):
    if extension == "xlsx":
        workbook = load_workbook(file_path)
        try:
            worksheet = workbook.active
            if keep_indexes is not None:
                keep_set = set(keep_indexes)
                for column_index in range(worksheet.max_column, 0, -1):
                    if column_index - 1 not in keep_set:
                        worksheet.delete_cols(column_index)
            for index, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=index).value = header
            workbook.save(file_path)
        finally:
            workbook.close()
        return file_path

    dataframe = pd.read_excel(file_path)
    if keep_indexes is not None:
        dataframe = dataframe.iloc[:, keep_indexes]
    dataframe.columns = headers
    converted_path = f"{os.path.splitext(file_path)[0]}.xlsx"
    dataframe.to_excel(converted_path, index=False)
    os.remove(file_path)
    return converted_path


def validate_and_normalize_dashboard_upload(
    file_path,
    extension,
    input_scheme,
    connection_factory=None,
    extra_fields=None,
):
    """Normalize supported annual-report headers to sequence + Chinese field name."""
    if input_scheme not in INPUT_SCHEMES:
        raise ValueError("請選擇正確的輸入欄位命名來源。")
    headers = _read_headers(file_path, extension)
    if not headers:
        raise ValueError("Excel 第一列沒有可辨識的欄位表頭。")
    rows = _load_field_map(connection_factory)
    normalized_headers = _normalize_headers(headers, input_scheme, rows)
    keep_indexes = None
    if extra_fields is not None:
        retained_extra_fields = {_text(field) for field in extra_fields}
        canonical_headers = set(_maps_for_scheme(rows, input_scheme)[0].values())
        keep_indexes = [
            index
            for index, (original, normalized) in enumerate(zip(headers, normalized_headers))
            if original != normalized or normalized in canonical_headers or original in retained_extra_fields
        ]
        normalized_headers = [normalized_headers[index] for index in keep_indexes]
        if not normalized_headers:
            raise ValueError("沒有可匯入的欄位，請至少保留一個欄位。")
    return _write_headers(file_path, extension, normalized_headers, keep_indexes)


def preview_dashboard_upload(file_source, extension, input_scheme, connection_factory=None):
    """Return a read-only header matching preview without storing or rewriting the file."""
    if input_scheme not in INPUT_SCHEMES:
        raise ValueError("請選擇正確的輸入欄位命名來源。")
    headers = _read_headers(file_source, extension)
    if not headers:
        raise ValueError("Excel 第一列沒有可辨識的欄位表頭。")
    rows = _load_field_map(connection_factory)
    normalized_headers = _normalize_headers(headers, input_scheme, rows)
    canonical_headers = set(_maps_for_scheme(rows, input_scheme)[0].values())
    columns = []
    for original, normalized in zip(headers, normalized_headers):
        columns.append({
            "original": original,
            "normalized": normalized,
            "matched": original != normalized or normalized in canonical_headers,
        })
    matched_count = sum(1 for column in columns if column["matched"])
    return {
        "columns": columns,
        "total_count": len(columns),
        "matched_count": matched_count,
        "unmatched_count": len(columns) - matched_count,
    }
