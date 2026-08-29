import logging
from datetime import datetime
from uuid import UUID, uuid4
from zipfile import BadZipFile
from flask import flash, redirect, render_template, request, session, url_for
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from modules.blueprint.auth.key_access import data_update_key_required
from modules.services.auth import login_required
from modules.services.dashboard import dashboard_bp
from modules.services.db import get_conn
from modules.blueprint.dashboard.branch_versions import (commit_changes, create_empty_commit_file, discard_staging_changes, ensure_initial_commit, finalize_staging_commit, has_any_staging_changes, record_change, replace_staging_changes, reverse_commits, save_staging_commit, version_changes)

NATIONAL_IMPORT_SHEETS = {
    "年齡": {"table": "National_Age", "item_column": "Age", "headers": ("年度", "癌別", "年齡", "合計", "男性", "女性", "醫學中心", "非醫學中心")},
    "整併期別": {"table": "National_Period", "item_column": "Period", "headers": ("年度", "癌別", "整併期別", "合計", "男性", "女性", "醫學中心", "非醫學中心")},}

NATIONAL_DATASETS = {
    "age": {"table": "National_Age", "id_column": "NationAge_ID", "item_column": "Age", "label": "年齡", "item_label": "年齡"},
    "period": {"table": "National_Period", "id_column": "NationPeriod_ID", "item_column": "Period", "label": "整併期別", "item_label": "整併期別"},}

NATIONAL_COLUMN_LABELS = {
    "Year": "年度", "Cancer": "癌別", "Total": "合計", "Male": "男性", "Female": "女性",
    "Medical_Centers": "醫學中心", "N_Medical_Centers": "非醫學中心",}

def read_national_import_rows(file_source):
    try:
        workbook = load_workbook(file_source, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as error:
        raise ValueError("無法讀取 Excel 檔案。") from error
    try:
        missing = [name for name in NATIONAL_IMPORT_SHEETS if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"Excel 缺少工作表：{'、'.join(missing)}。")
        imported_rows = {}
        for sheet_name, config in NATIONAL_IMPORT_SHEETS.items():
            worksheet = workbook[sheet_name]
            try:
                headers = tuple(cell.value for cell in next(worksheet.iter_rows(max_row=1)))
            except StopIteration as error:
                raise ValueError(f"「{sheet_name}」工作表沒有標題列。") from error
            if headers != config["headers"]:
                raise ValueError(f"「{sheet_name}」工作表的欄位順序或名稱不正確。")
            rows = []
            for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
                    continue
               
                year, cancer, item, *counts = values
                if any(isinstance(value, bool) for value in counts):
                    raise ValueError
                
                counts = [int(value) for value in counts]
                if any(value < 0 for value in counts):
                    raise ValueError(f"「{sheet_name}」工作表第 {row_number} 列的數值欄位不可為負數。")
                rows.append((
                    str(int(year)) if isinstance(year, float) and year.is_integer() else str(year).strip(),
                    str(cancer).strip(),
                    str(int(item)) if isinstance(item, float) and item.is_integer() else str(item).strip(),
                    *counts,
                ))
            imported_rows[sheet_name] = rows
        return imported_rows
    finally:
        workbook.close()

def national_rows(dataset, user_id):
    config = NATIONAL_DATASETS[dataset]
    fields = ["Year", "Cancer", config["item_column"], "Total", "Male", "Female", "Medical_Centers", "N_Medical_Centers"]
    conn = get_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT " + ", ".join(f"[{field}]" for field in [config["id_column"], *fields]) + f" FROM dbo.[{config['table']}]")
        rows = {str(row[0]): tuple(row[1:]) for row in cursor.fetchall()}
    finally:
        conn.close()
    for change in version_changes("national", user_id):
        if change.get("dataset") != dataset:
            continue
        record_id = str(change["record_id"])
        if change["action"] == "Delete":
            rows.pop(record_id, None)
        else:
            rows[record_id] = tuple(change["after"].get(field) for field in fields)
    return fields, rows


def get_national_data(dataset, user_id, selected_year="", search_column="", search_query=""):
    config = NATIONAL_DATASETS[dataset]
    fields, all_rows = national_rows(dataset, user_id)
    year_counts = {}
    for row in all_rows.values():
        if row[0] is not None:
            year = str(row[0])
            year_counts[year] = year_counts.get(year, 0) + 1
    years = sorted(year_counts, key=lambda year: (year.isdigit(), int(year) if year.isdigit() else year), reverse=True)
    selected_year = selected_year if selected_year in years else next(iter(years), "")
    search_column = search_column if search_column in {"Cancer", config["item_column"]} else ""
    search_index = fields.index(search_column or "Cancer")
    search_query = search_query.strip()
    rows = [(record_id, *row) for record_id, row in all_rows.items() if (not selected_year or str(row[0]) == selected_year) and (not search_query or search_query.lower() in str(row[search_index] or "").lower())]
    rows.sort(key=lambda row: (str(row[2] or ""), str(row[3] or "")))
    return {"config": config, "rows": rows, "available_years": years, "selected_year": selected_year, "selected_year_count": year_counts.get(selected_year, 0), "search_column": search_column, "search_query": search_query}

def insert_national_rows(imported_rows):
    conn = get_conn()
    try:
        cursor = conn.cursor()
        for sheet_name, config in NATIONAL_IMPORT_SHEETS.items():
            for row in imported_rows[sheet_name]:
                record_change(cursor, "national", session["userid"], f"draft-{uuid4()}", "Create", after=dict(zip(["Year", "Cancer", config["item_column"], "Total", "Male", "Female", "Medical_Centers", "N_Medical_Centers"], row)), dataset='age' if sheet_name == '年齡' else 'period')
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def delete_national_rows(dataset, record_ids):
    config = NATIONAL_DATASETS[dataset]
    fields, rows = national_rows(dataset, session["userid"])
    conn = get_conn()
    try:
        cursor = conn.cursor()
        deleted_count = 0
        for record_id in record_ids:
            previous = rows.get(str(record_id))
            if previous is not None:
                record_change(cursor, "national", session["userid"], record_id, "Delete", before=dict(zip(fields, previous)), dataset=dataset)
                deleted_count += 1
        return deleted_count
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def update_national_row(dataset, record_id, values):
    config = NATIONAL_DATASETS[dataset]
    fields, rows = national_rows(dataset, session["userid"])
    previous = rows.get(str(record_id))
    if previous is None:
        raise ValueError("找不到要修改的資料。")
    conn = get_conn()
    try:
        cursor = conn.cursor()
        record_change(cursor, "national", session["userid"], record_id, "Update", before=dict(zip(fields, previous)), after=dict(zip(fields, values)), dataset=dataset)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def redirect_national_import(draft=False):
    return redirect(url_for("dashboard.national_import", view=request.form.get("view", "age"), year=request.form.get("year", ""), column=request.form.get("column", ""), q=request.form.get("q", ""), draft="1" if draft else None))


def apply_national_changes(cursor, changes):
    draft_ids = {}
    for change in changes:
        config = NATIONAL_DATASETS[change["dataset"]]
        fields = ["Year", "Cancer", config["item_column"], "Total", "Male", "Female", "Medical_Centers", "N_Medical_Centers"]
        record_id = draft_ids.get(str(change["record_id"]), change["record_id"])
        if change["action"] == "Create":
            cursor.execute(f"INSERT INTO dbo.[{config['table']}] ({', '.join(f'[{field}]' for field in fields)}) OUTPUT INSERTED.[{config['id_column']}] VALUES ({', '.join('?' for _ in fields)})", *(change["after"].get(field) for field in fields))
            actual_id = cursor.fetchone()[0]
            draft_ids[str(change["record_id"])] = actual_id
            change["record_id"] = actual_id
        elif change["action"] == "Update":
            cursor.execute(f"UPDATE dbo.[{config['table']}] SET {', '.join(f'[{field}] = ?' for field in fields)} WHERE [{config['id_column']}] = ?", *(change["after"].get(field) for field in fields), record_id)
            change["record_id"] = record_id
        else:
            cursor.execute(f"DELETE FROM dbo.[{config['table']}] WHERE [{config['id_column']}] = ?", record_id)
            change["record_id"] = record_id


@dashboard_bp.route("/dashboard/national-import/versions/save", methods=["POST"])
@login_required
@data_update_key_required
def save_national_import_version():
    conn = get_conn()
    try:
        cursor = conn.cursor()
        changes = version_changes("national", session["userid"])
        apply_national_changes(cursor, changes)
        commit_id, _ = save_staging_commit(cursor, "national", session["userid"], session.get("name") or session["userid"], request.form.get("info", ""), changes)
        conn.commit()
        replace_staging_changes("national", session["userid"], changes)
        finalize_staging_commit("national", session["userid"], commit_id)
        flash(f"已儲存國家資料版本 {commit_id}。", "success")
    except ValueError as error:
        conn.rollback()
        flash(str(error), "warning")
    except Exception:
        conn.rollback()
        logging.exception("Unable to save national version")
        flash("儲存國家資料版本失敗。", "danger")
    finally:
        conn.close()
    return redirect_national_import()

@dashboard_bp.route("/dashboard/national-import/versions")
@login_required
@data_update_key_required
def national_import_versions():
    conn = get_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT CommitId, ParentCommitId, CreatedByUserID, CreatedByName, Info, CreatedAt, Action, ChangeCount, RevertUntil FROM dbo.BranchCommits WHERE DataType = 'national' ORDER BY CreatedAt DESC")
        versions = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    finally:
        conn.close()
    return render_template("branch_version.html", active="data_update_access", title="全國資料版本紀錄", description="年齡與整併期別會一起儲存及回復。", back_url=url_for("dashboard.national_import"), version_endpoint="dashboard.national_import_versions", preview_endpoint="dashboard.preview_national_import_version", restore_endpoint="dashboard.restore_national_import_version", versions=versions, now=datetime.now(), show_filters=False)


@dashboard_bp.route("/dashboard/national-import/versions/<version_id>/preview")
@login_required
@data_update_key_required
def preview_national_import_version(version_id):
    try:
        target_id = UUID(version_id)
    except ValueError:
        flash("版本編號無效。", "danger")
        return redirect(url_for("dashboard.national_import_versions"))
    if has_any_staging_changes("national"):
        flash("有未儲存變更時無法預覽歷史版本。", "danger")
        return redirect(url_for("dashboard.national_import_versions"))

    conn = get_conn()
    try:
        cursor = conn.cursor()
        commits = reverse_commits(cursor, "national", target_id)
        if commits is None:
            flash("找不到指定版本。", "danger")
            return redirect(url_for("dashboard.national_import_versions"))

        states = {}
        for key, config in NATIONAL_DATASETS.items():
            fields = ["Year", "Cancer", config["item_column"], "Total", "Male", "Female", "Medical_Centers", "N_Medical_Centers"]
            cursor.execute(f"SELECT [{config['id_column']}], {', '.join(f'[{field}]' for field in fields)} FROM dbo.[{config['table']}]")
            states[key] = {row[0]: dict(zip(fields, row[1:])) for row in cursor.fetchall()}

        for commit_id in commits:
            for change in reversed(commit_changes("national", commit_id)):
                state = states[change["dataset"]]
                if change["action"] == "Create":
                    state.pop(change["record_id"], None)
                else:
                    state[change["record_id"]] = change["before"]
    finally:
        conn.close()

    dataset = request.args.get("view", "age")
    dataset = dataset if dataset in NATIONAL_DATASETS else "age"
    config = NATIONAL_DATASETS[dataset]
    fields = ["Year", "Cancer", config["item_column"], "Total", "Male", "Female", "Medical_Centers", "N_Medical_Centers"]
    year_counts = {}
    for row in states[dataset].values():
        if row.get("Year") is not None:
            year = str(row["Year"])
            year_counts[year] = year_counts.get(year, 0) + 1
    available_years = sorted(year_counts, key=lambda year: (year.isdigit(), int(year) if year.isdigit() else year), reverse=True)
    selected_year = request.args.get("year", "")
    selected_year = selected_year if selected_year in available_years else next(iter(available_years), "")
    search_column = request.args.get("column", "")
    search_column = search_column if search_column in {"Cancer", config["item_column"]} else ""
    search_query = request.args.get("q", "").strip()
    rows = [(record_id, *(row.get(field) for field in fields)) for record_id, row in states[dataset].items() if (not selected_year or str(row.get("Year")) == selected_year) and (not search_query or search_query.lower() in str(row.get(search_column or "Cancer") or "").lower())]
    rows.sort(key=lambda row: (str(row[2] or ""), str(row[3] or "")))
    return render_template("national_import.html", active="data_update_access", version_id=version_id, preview_mode=True, preview_commit_id=version_id, pending_change_count=0, dataset=dataset, datasets=NATIONAL_DATASETS, config=config, column_labels=NATIONAL_COLUMN_LABELS, rows=rows, available_years=available_years, selected_year=selected_year, selected_year_count=year_counts.get(selected_year, 0), search_column=search_column, search_query=search_query, editing_id="")


@dashboard_bp.route("/dashboard/national-import/versions/<version_id>/restore", methods=["POST"])
@login_required
@data_update_key_required
def restore_national_import_version(version_id):
    try:
        target_id = UUID(version_id)
    except ValueError:
        flash("版本編號無效。", "danger")
        return redirect(url_for("dashboard.national_import_versions"))
    if has_any_staging_changes("national"):
        flash("仍有未儲存變更，無法回復版本。", "danger")
        return redirect(url_for("dashboard.national_import_versions"))

    conn = get_conn()
    try:
        cursor = conn.cursor()
        commits = reverse_commits(cursor, "national", target_id)
        if commits is None:
            raise ValueError("找不到指定版本。")
        changes = [change for commit_id in commits for change in reversed(commit_changes("national", commit_id))]
        for change in changes:
            config = NATIONAL_DATASETS[change["dataset"]]
            fields = ["Year", "Cancer", config["item_column"], "Total", "Male", "Female", "Medical_Centers", "N_Medical_Centers"]
            cursor.execute(f"SELECT {', '.join(f'[{field}]' for field in fields)} FROM dbo.[{config['table']}] WHERE [{config['id_column']}] = ?", change["record_id"])
            current = cursor.fetchone()
            after = tuple(change["after"].get(field) for field in fields)
            if (change["action"] in {"Create", "Update"} and (not current or tuple(current) != after)) or (change["action"] == "Delete" and current):
                raise ValueError("資料已被其他使用者修改，無法回復版本。")

        for change in changes:
            before = change["before"]
            after = change["after"]
            if change["action"] == "Create":
                record_change(cursor, "national", session["userid"], change["record_id"], "Delete", before=after, dataset=change["dataset"])
            elif change["action"] == "Update":
                record_change(cursor, "national", session["userid"], change["record_id"], "Update", before=after, after=before, dataset=change["dataset"])
            else:
                record_change(cursor, "national", session["userid"], change["record_id"], "Create", after=before, dataset=change["dataset"])
        flash(f"已回復至版本 {version_id}，請儲存目前版本完成提交。", "success")
    except ValueError as error:
        conn.rollback()
        flash(str(error), "danger")
    except Exception:
        conn.rollback()
        logging.exception("Unable to restore national version")
        flash("回復國家資料版本失敗。", "danger")
    finally:
        conn.close()
    return redirect(url_for("dashboard.national_import", draft="1"))

@dashboard_bp.route("/dashboard/national-import", methods=["GET", "POST"])
@login_required
@data_update_key_required
def national_import():
    if request.method == "GET":
        dataset = request.args.get("view", "age")
        dataset = dataset if dataset in NATIONAL_DATASETS else "age"
        if request.args.get("draft") != "1":
            discard_staging_changes("national", session["userid"])
        conn = get_conn()
        try:
            commit_id = ensure_initial_commit(conn.cursor(), "national", session["userid"], session.get("name") or session["userid"])
            conn.commit()
            if commit_id:
                create_empty_commit_file("national", commit_id)
        finally:
            conn.close()
        try:
            data = get_national_data(dataset, session["userid"], request.args.get("year", ""), request.args.get("column", ""), request.args.get("q", ""))
        except Exception:
            logging.exception("Unable to read national import data")
            flash("無法讀取資料表。", "danger")
            data = {"config": NATIONAL_DATASETS[dataset], "rows": [], "available_years": [], "selected_year": "", "selected_year_count": 0, "search_column": "", "search_query": ""}
        return render_template("national_import.html", active="data_update_access", dataset=dataset, datasets=NATIONAL_DATASETS, column_labels=NATIONAL_COLUMN_LABELS, editing_id=request.args.get("edit_id", "").strip(), pending_change_count=len(version_changes("national", session["userid"])), **data)

    uploaded_file = request.files.get("import_file")
    try:
        if not uploaded_file or not uploaded_file.filename:
            raise ValueError("請選擇國家資料 Excel 檔案。")
        if not uploaded_file.filename.lower().endswith(".xlsx"):
            raise ValueError("僅接受 .xlsx 格式的國家資料檔案。")
        imported_rows = read_national_import_rows(uploaded_file)
        insert_national_rows(imported_rows)
    except ValueError as error:
        flash(str(error), "danger")
    except Exception:
        logging.exception("Unable to import national data")
        flash("匯入失敗，資料未寫入。請確認資料庫連線與資料表。", "danger")
    else:
        flash(f"匯入完成：National_Age {len(imported_rows['年齡'])} 筆、National_Period {len(imported_rows['整併期別'])} 筆。", "success")
    return redirect_national_import(draft=True)

@dashboard_bp.route("/dashboard/national-import/delete", methods=["POST"])
@login_required
@data_update_key_required
def delete_national_rows_route():
    dataset, record_ids = request.form.get("view", "age"), request.form.getlist("record_ids")
    try:
        if dataset not in NATIONAL_DATASETS:
            raise ValueError("資料類型不正確。")
        if not record_ids:
            raise ValueError("請先選擇要刪除的資料。")
        if any(not (record_id.isdigit() or record_id.startswith("draft-")) for record_id in record_ids):
            raise ValueError("刪除資料識別碼不正確。")
        deleted_count = delete_national_rows(dataset, record_ids)
    except ValueError as error:
        flash(str(error), "warning")
    except Exception:
        logging.exception("Unable to delete national data")
        flash("刪除失敗，資料未變更。", "danger")
    else:
        flash(f"已刪除 {deleted_count} 筆{NATIONAL_DATASETS[dataset]['label']}資料。", "success")
    return redirect_national_import(draft=True)


@dashboard_bp.route("/dashboard/national-import/<record_id>/update", methods=["POST"])
@login_required
@data_update_key_required
def update_national_row_route(record_id):
    dataset = request.form.get("view", "age")
    config = NATIONAL_DATASETS.get(dataset)
    values = [request.form.get(name, "").strip() for name in ("Year", "Cancer", config["item_column"] if config else "", "Total", "Male", "Female", "Medical_Centers", "N_Medical_Centers")]
    try:
        if not config:
            raise ValueError("資料類型不正確。")
        if not all(values[:3]):
            raise ValueError("年度、癌別與資料項目不可空白。")
        if any(not value.isdigit() for value in values[3:]):
            raise ValueError("合計、男性、女性、醫學中心與非醫學中心必須為非負整數。")
        update_national_row(dataset, record_id, (*values[:3], *(int(value) for value in values[3:])))
    except ValueError as error:
        flash(str(error), "danger")
    except Exception:
        logging.exception("Unable to update national data")
        flash("修改資料時發生錯誤。", "danger")
    else:
        flash("資料已修改。", "success")
    return redirect_national_import(draft=True)


@dashboard_bp.route("/dashboard/national-import/<record_id>/delete", methods=["POST"])
@login_required
@data_update_key_required
def delete_national_row_route(record_id):
    dataset = request.form.get("view", "age")
    try:
        if dataset not in NATIONAL_DATASETS:
            raise ValueError("資料類型不正確。")
        deleted_count = delete_national_rows(dataset, [str(record_id)])
    except ValueError as error:
        flash(str(error), "danger")
    except Exception:
        logging.exception("Unable to delete national data")
        flash("刪除資料時發生錯誤。", "danger")
    else:
        flash(f"已刪除 {deleted_count} 筆資料。", "success")
    return redirect_national_import(draft=True)
