from zipfile import BadZipFile
from functools import wraps
import secrets
from datetime import datetime
from uuid import UUID
from flask import Blueprint, flash, redirect, render_template, request, session, url_for
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from modules.services.auth import login_required
from modules.services.db import get_conn
from modules.blueprint.dashboard.histology_versions import (commit_ids_equal,commit_changes,create_empty_commit_file,ensure_initial_commit,finalize_staging_commit,has_any_staging_changes,latest_commit_id,mapping_values_match,normalize_change,record_staging_change,save_staging_commit,staging_count,)

histology_mapping_bp = Blueprint("histology_mapping",__name__,template_folder="templates",)

_FIELD_CONFIG = (
    ("CodeYear", "年度", ("codeyear", "code_year"), True, True),
    ("CancerGroupKey", "癌別群組", ("cancergroupkey", "cancer_group_key"), False, True),
    ("CancerGroup_zh", "癌別(中)", ("cancergroup_zh",), True, True),
    ("CancerGroup_en", "癌別(英)", ("cancergroup_en",), True, True),
    ("hist", "性態碼", ("hist", "histcode"), True, True),
    ("behavior", "行為碼", ("behavior", "behaviorcode"), True, True),
    ("hist_zh", "組織型態(中)", ("hist_zh", "histologyzh"), True, True),
    ("hist_en", "組織型態(英)", ("hist_en", "histologyen", "behavior_en", "histology"), True, True),
    ("SiteInclude", "納入部位", ("siteinclude", "site_include"), False, False),
    ("SiteExclude", "排除部位", ("siteexclude", "site_exclude"), False, False),
)

_COLUMN_LABELS = {
    alias: label
    for _, label, aliases, _, _ in _FIELD_CONFIG
    for alias in aliases
}

_DISPLAY_COLUMN_NAMES = tuple(
    aliases[0]
    for _, _, aliases, show_in_table, _ in _FIELD_CONFIG
    if show_in_table
)

_IMPORT_COLUMNS = tuple(
    column
    for column, _, _, _, allow_excel_import in _FIELD_CONFIG
    if allow_excel_import
)

_NUMERIC_COLUMNS = frozenset({"codeyear", "code_year", "hist", "histcode", "behavior", "behaviorcode"})
_HISTOLOGY_ACCESS_TOKEN = "histology_mapping_access_token"
_HISTOLOGY_ACCESS_USER = "histology_mapping_access_user"

def _has_histology_access(access_token):
    return session.get("position") == "Admin" or (bool(access_token) and session.get(_HISTOLOGY_ACCESS_USER) == session.get("userid") and secrets.compare_digest(access_token, session.get(_HISTOLOGY_ACCESS_TOKEN, "")))

def histology_key_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not _has_histology_access(request.form.get("access_token", "")):
            flash("請先輸入有效且已啟用的 Key，才能修改組織型態資料。", "danger")
            return redirect(url_for("histology_mapping.histology_code_mapping"))
        return view(*args, **kwargs)
    return wrapped_view

def _quote_identifier(identifier):
    return f"[{identifier.replace(']', ']]')}]"

def _get_mapping_columns():
    conn = get_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 0 * FROM dbo.histology_code_mapping")
        return [column[0] for column in cursor.description]
    finally:
        conn.close()

def _get_mapping_data(search_column="", search_query="", selected_year=""):
    conn = get_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 0 * FROM dbo.histology_code_mapping")
        columns = [column[0] for column in cursor.description]
        query = "SELECT * FROM dbo.histology_code_mapping"
        conditions = []
        parameters = []
        code_year_column = next((column for column in columns if column.lower() in {"codeyear", "code_year"}),None,)
        if selected_year and code_year_column:
            conditions.append(f"{_quote_identifier(code_year_column)} = ?")
            parameters.append(selected_year)
        search_query = search_query.strip()
        if search_query:
            searchable_columns = [search_column] if search_column in columns else columns
            expression = " + N' ' + ".join(
                f"COALESCE(CONVERT(NVARCHAR(MAX), {_quote_identifier(column)}), N'')"
                for column in searchable_columns
            )
            conditions.append(f"({expression}) LIKE ?")
            parameters.append(f"%{search_query}%")
        if conditions:
            query += f" WHERE {' AND '.join(conditions)}"
        cursor.execute(query, *parameters)
        return columns, [tuple(row) for row in cursor.fetchall()]
    finally:
        conn.close()

def _get_year_counts():
    conn = get_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 0 * FROM dbo.histology_code_mapping")
        columns = [column[0] for column in cursor.description]
        code_year_column = next((column for column in columns if column.lower() in {"codeyear", "code_year"}),None,)
        if not code_year_column:
            return columns, []

        quoted_year_column = _quote_identifier(code_year_column)
        cursor.execute(f"SELECT {quoted_year_column}, COUNT(*) FROM dbo.histology_code_mapping WHERE {quoted_year_column} IS NOT NULL GROUP BY {quoted_year_column}")
        year_counts = [(str(row[0]), row[1]) for row in cursor.fetchall()]
        year_counts.sort(key=lambda item: (item[0].isdigit(),int(item[0]) if item[0].isdigit() else item[0],),reverse=True,)
        return columns, year_counts
    finally:
        conn.close()


def _primary_key(columns):
    return next((column for column in columns if column.lower() == "histcode_id"), None)

def _return_to_mapping():
    return redirect(
        url_for(
            "histology_mapping.histology_code_mapping",
            year=request.form.get("return_year", ""),
            column=request.form.get("return_column", ""),
            q=request.form.get("return_q", ""),
            access_token=request.form.get("access_token", ""),
        ))

def _editable_columns(columns):
    primary_key = _primary_key(columns)
    return [column for column in columns if column != primary_key]


def _validated_mapping_values(columns, values):
    values = ["" if value is None else str(value).strip() for value in values]
    empty_columns = [
        _COLUMN_LABELS.get(column.lower(), column)
        for column, value in zip(columns, values)
        if not value
        ]
    if empty_columns:
        raise ValueError(f"請完整填寫欄位：{'、'.join(empty_columns)}。")

    invalid_columns = [
        _COLUMN_LABELS.get(column.lower(), column)
        for column, value in zip(columns, values)
        if column.lower() in _NUMERIC_COLUMNS and not value.isdigit()]
    
    if invalid_columns:
        raise ValueError(f"{'、'.join(invalid_columns)} 僅能輸入整數。")
    return values

def _insert_mapping_rows(columns, rows):
    editable_columns = _editable_columns(columns)
    primary_key = _primary_key(columns)
    fields = ", ".join(_quote_identifier(column) for column in editable_columns)
    placeholders = ", ".join("?" for _ in editable_columns)
    conn = get_conn()
    try:
        cursor = conn.cursor()
        for row in rows:
            cursor.execute(f"INSERT INTO dbo.histology_code_mapping ({fields}) OUTPUT INSERTED.{_quote_identifier(primary_key)} VALUES ({placeholders})",*row,)
            record_staging_change(cursor, session["userid"], cursor.fetchone()[0], "Create", after=dict(zip(editable_columns, row)))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def _read_import_rows(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = tuple(cell.value for cell in next(worksheet.iter_rows(max_row=1)))
        if headers != _IMPORT_COLUMNS:
            raise ValueError("Excel 欄位須完全符合組織型態表範本。")

        rows = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value is None or str(value).strip() == "" for value in values):
                continue
            if any(value is None or str(value).strip() == "" for value in values):
                raise ValueError(f"第 {row_number} 列有未填寫欄位。")
            rows.append(tuple(str(value).strip() for value in values))
        if not rows:
            raise ValueError("Excel 沒有可新增的資料。")
        return rows
    finally:
        workbook.close()

@histology_mapping_bp.route("/dashboard/histology-code")
@login_required
def histology_code_mapping():
    search_column = request.args.get("column", "")
    search_query = request.args.get("q", "")
    access_token = request.args.get("access_token", "")
    if not _has_histology_access(access_token):
        session.pop(_HISTOLOGY_ACCESS_TOKEN, None)
        session.pop(_HISTOLOGY_ACCESS_USER, None)
        return render_template("histology_key_access.html",active="histology_code_mapping",return_year=request.args.get("year", ""),return_column=search_column,return_q=search_query,)

    edit_id = request.args.get("edit_id", "")
    columns, year_counts = _get_year_counts()
    available_years = [year for year, _ in year_counts]
    requested_year = request.args.get("year", "")
    selected_year = requested_year if requested_year in available_years else next(iter(available_years), "")
    columns, rows = _get_mapping_data(search_column, search_query, selected_year)
    selected_year_count = dict(year_counts).get(selected_year, 0)
    primary_key = _primary_key(columns)
    editing_row = None
    if edit_id and primary_key:
        key_index = columns.index(primary_key)
        editing_row = next((row for row in rows if str(row[key_index]) == edit_id), None)

    conn = get_conn()
    cursor = conn.cursor()
    initial_commit_id = ensure_initial_commit(cursor, session["userid"], session.get("name"))
    if initial_commit_id:
        conn.commit()
        create_empty_commit_file(initial_commit_id)
    pending_change_count = staging_count(session["userid"])
    conn.close()

    return render_template("histology_code.html",active="histology_code_mapping",columns=columns,
        display_columns=[column for name in _DISPLAY_COLUMN_NAMES for column in columns if column.lower() == name],
        column_labels={column: _COLUMN_LABELS.get(column.lower(), column) for column in columns},
        rows=rows,
        available_years=available_years,
        selected_year=selected_year,
        selected_year_count=selected_year_count,
        primary_key=primary_key,
        editable_columns=_editable_columns(columns),
        search_column=search_column if search_column in columns else "",
        search_query=search_query,
        editing_row=editing_row,
        access_token=access_token,
        numeric_columns=_NUMERIC_COLUMNS,
        pending_change_count=pending_change_count,
    )

@histology_mapping_bp.route("/dashboard/histology-code/access", methods=["POST"])
@login_required
def verify_histology_mapping_access():
    api_key = request.form.get("api_key", "").strip()
    if not api_key:
        flash("請輸入金鑰。", "danger")
        return redirect(url_for("histology_mapping.histology_code_mapping"))

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("""SELECT TOP 1 1 FROM dbo.User_applications WHERE UserID = ? AND API_key = ? AND Status = 'Active' AND End_time >= GETDATE()""",session["userid"],api_key,)
    key_is_valid = cursor.fetchone() is not None
    conn.close()

    if not key_is_valid:
        flash("金鑰失效、未啟用或帳號不符，導致啟用失敗，請確認您的金鑰與登入帳號是否正確", "danger")
        return redirect(url_for("histology_mapping.histology_code_mapping"))

    access_token = secrets.token_urlsafe(24)
    session[_HISTOLOGY_ACCESS_TOKEN] = access_token
    session[_HISTOLOGY_ACCESS_USER] = session["userid"]
    return redirect(url_for("histology_mapping.histology_code_mapping",year=request.form.get("return_year", ""),column=request.form.get("return_column", ""),q=request.form.get("return_q", ""),access_token=access_token,))

@histology_mapping_bp.route("/dashboard/histology-code/versions")
@login_required
def histology_code_versions():
    access_token = request.args.get("access_token", "")
    if not _has_histology_access(access_token):
        return redirect(url_for("histology_mapping.histology_code_mapping"))

    selected_user_id = request.args.get("user_id", "")
    selected_date = request.args.get("date", "")

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT CommitId, CommitId AS VersionId, ParentCommitId, CreatedByUserID, CreatedByName, CreatedAt, Action, ChangeCount, ChangeCount AS RecordCount, RevertUntil FROM dbo.HistologyCodeCommits ORDER BY CreatedAt DESC")
    versions = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    users = []
    seen_user_ids = set()
    for version in versions:
        user_id = str(version["CreatedByUserID"] or "")
        if user_id and user_id not in seen_user_ids:
            users.append({"id": user_id,"name": version["CreatedByName"] or user_id,})
            seen_user_ids.add(user_id)

    if selected_user_id:
        versions = [version
            for version in versions
            if str(version["CreatedByUserID"] or "") == selected_user_id]
    if selected_date:
        versions = [version
            for version in versions
            if version["CreatedAt"] and version["CreatedAt"].date().isoformat() == selected_date]
    conn.close()
    return render_template("histology_code_version.html",active="histology_code_mapping",versions=versions,users=users,access_token=access_token,selected_user_id=selected_user_id,selected_date=selected_date,now=datetime.now(),)

@histology_mapping_bp.route("/dashboard/histology-code/versions/save", methods=["POST"])
@login_required
@histology_key_required
def save_histology_code_version():
    conn = get_conn()
    try:
        cursor = conn.cursor()
        commit_id, _ = save_staging_commit(cursor, session["userid"], session.get("name"))
        conn.commit()
        finalize_staging_commit(session["userid"], commit_id)
        flash(f"已儲存版本 {commit_id}。", "success")
        return _return_to_mapping()
    finally:
        conn.close()

@histology_mapping_bp.route("/dashboard/histology-code/versions/<version_id>/restore", methods=["POST"])
@login_required
@histology_key_required
def restore_histology_code_version(version_id):
    return restore_histology_code_timeline(version_id)

@histology_mapping_bp.route("/dashboard/histology-code/versions/<version_id>/restore-timeline", methods=["POST"])
@login_required
@histology_key_required
def restore_histology_code_timeline(version_id):
    try:
        target_id = UUID(version_id)
    except ValueError:
        flash("版本識別碼無效。", "danger")
        return redirect(url_for("histology_mapping.histology_code_versions", access_token=request.form.get("access_token", "")))
    conn = get_conn()
    try:
        cursor = conn.cursor()
        if has_any_staging_changes():
            flash("無法回到此時間點：仍有未儲存變更。", "danger")
            return redirect(url_for("histology_mapping.histology_code_versions", access_token=request.form.get("access_token", "")))
        reverse_commits = []
        current_id = latest_commit_id(cursor)
        while current_id and not commit_ids_equal(current_id, target_id):
            cursor.execute("SELECT ParentCommitId FROM dbo.HistologyCodeCommits WHERE CommitId = ?", current_id)
            parent = cursor.fetchone()
            if not parent:
                break
            reverse_commits.append(current_id)
            current_id = parent[0]
        if not commit_ids_equal(current_id, target_id):
            flash("選取版本不在目前版本紀錄。", "danger")
            return redirect(url_for("histology_mapping.histology_code_versions", access_token=request.form.get("access_token", "")))
        columns = _get_mapping_columns()
        primary_key, editable_columns = _primary_key(columns), _editable_columns(columns)
        changes = []
        for commit_id in reverse_commits:
            changes.extend(reversed(commit_changes(commit_id)))
        changes = [normalize_change(change, editable_columns) for change in changes]
        for change in changes:
            cursor.execute(f"SELECT {', '.join(_quote_identifier(column) for column in editable_columns)} FROM dbo.histology_code_mapping WHERE {_quote_identifier(primary_key)} = ?", change["HistCodeId"])
            current = cursor.fetchone()
            after = {column: change.get(f"After{column}") for column in editable_columns}
            if (change["Action"] in ("Create", "Update") and (not current or not mapping_values_match(editable_columns, dict(zip(editable_columns, current)), after)) ) or (change["Action"] == "Delete" and current):
                flash("無法回到此時間點：資料已與版本紀錄不一致。", "danger")
                return redirect(url_for("histology_mapping.histology_code_versions", access_token=request.form.get("access_token", "")))
        for change in changes:
            histcode_id = change["HistCodeId"]
            before = {column: change.get(f"Before{column}") for column in editable_columns}
            after = {column: change.get(f"After{column}") for column in editable_columns}
            if change["Action"] == "Create":
                cursor.execute(f"DELETE FROM dbo.histology_code_mapping WHERE {_quote_identifier(primary_key)} = ?", histcode_id)
                record_staging_change(cursor, session["userid"], histcode_id, "Delete", before=after)
            elif change["Action"] == "Update":
                assignments = ", ".join(f"{_quote_identifier(column)} = ?" for column in editable_columns)
                cursor.execute(f"UPDATE dbo.histology_code_mapping SET {assignments} WHERE {_quote_identifier(primary_key)} = ?", *before.values(), histcode_id)
                record_staging_change(cursor, session["userid"], histcode_id, "Update", before=after, after=before)
            else:
                fields = ", ".join(_quote_identifier(column) for column in editable_columns)
                cursor.execute(f"INSERT INTO dbo.histology_code_mapping ({fields}) VALUES ({', '.join('?' for _ in editable_columns)})", *before.values())
                record_staging_change(cursor, session["userid"], histcode_id, "Create", after=before)
        conn.commit()
        flash(f"已回到版本 {version_id} 的時間點；請確認或繼續修改後，再儲存目前版本。", "success")
    except Exception:
        conn.rollback()
        flash("回到指定時間點失敗，資料未變更。", "danger")
    finally:
        conn.close()
    return redirect(url_for("histology_mapping.histology_code_mapping", access_token=request.form.get("access_token", "")))


@histology_mapping_bp.route("/dashboard/histology-code/versions/<version_id>/preview")
@login_required
def preview_histology_code_version(version_id):
    access_token = request.args.get("access_token", "")
    if not _has_histology_access(access_token):
        return redirect(url_for("histology_mapping.histology_code_mapping"))
    if has_any_staging_changes():
        flash("有未儲存變更時無法預覽歷史版本。", "danger")
        return redirect(url_for("histology_mapping.histology_code_versions", access_token=access_token))
    try:
        target_id = UUID(version_id)
    except ValueError:
        flash("版本識別碼無效。", "danger")
        return redirect(url_for("histology_mapping.histology_code_versions", access_token=access_token))
    conn = get_conn()
    try:
        cursor = conn.cursor()
        reverse_commits, current_id = [], latest_commit_id(cursor)
        while current_id and not commit_ids_equal(current_id, target_id):
            cursor.execute("SELECT ParentCommitId FROM dbo.HistologyCodeCommits WHERE CommitId = ?", current_id)
            parent = cursor.fetchone()
            if not parent:
                break
            reverse_commits.append(current_id)
            current_id = parent[0]
        if not commit_ids_equal(current_id, target_id):
            flash("選取版本不在目前版本紀錄。", "danger")
            return redirect(url_for("histology_mapping.histology_code_versions", access_token=access_token))
    finally:
        conn.close()

    columns, rows = _get_mapping_data()
    primary_key = _primary_key(columns)
    editable_columns = _editable_columns(columns)
    key_index = columns.index(primary_key)
    state = {row[key_index]: dict(zip(columns, row)) for row in rows}
    for commit_id in reverse_commits:
        for raw_change in reversed(commit_changes(commit_id)):
            change = normalize_change(raw_change, editable_columns)
            histcode_id = change["HistCodeId"]
            if change["Action"] == "Create":
                state.pop(histcode_id, None)
            else:
                values = {
                    column: change[f"Before{column}"]
                    for column in editable_columns
                }
                values[primary_key] = histcode_id
                state[histcode_id] = values

    code_year_column = next((column for column in columns if column.lower() in {"codeyear", "code_year"}), None)
    available_years = []
    selected_year = ""
    if code_year_column:
        year_counts = {}
        for row in state.values():
            year_val = str(row.get(code_year_column, ""))
            if year_val:
                year_counts[year_val] = year_counts.get(year_val, 0) + 1
        available_years = sorted(year_counts.keys(), key=lambda x: int(x) if x.isdigit() else x, reverse=True)
        requested_year = request.args.get("year", "")
        selected_year = requested_year if requested_year in available_years else next(iter(available_years), "")
        if selected_year:
            state = {k: v for k, v in state.items() if str(v.get(code_year_column, "")) == selected_year}

    preview_rows = [tuple(row.get(column) for column in columns) for row in state.values()]

    search_column = request.args.get("column", "")
    search_query = request.args.get("q", "").strip()
    if search_query:
        filtered_rows = []
        for row in preview_rows:
            if search_column and search_column in columns:
                val = str(row[columns.index(search_column)]) if row[columns.index(search_column)] is not None else ""
                if search_query.lower() in val.lower():
                    filtered_rows.append(row)
            else:
                if any(search_query.lower() in (str(val).lower() if val is not None else "") for val in row):
                    filtered_rows.append(row)
        preview_rows = filtered_rows

    return render_template("histology_code.html",active="histology_code_mapping",columns=columns,
        display_columns=[column for name in _DISPLAY_COLUMN_NAMES for column in columns if column.lower() == name],
        column_labels={column: _COLUMN_LABELS.get(column.lower(), column) for column in columns},
        rows=preview_rows,
        available_years=available_years,
        selected_year=selected_year,
        selected_year_count=len(preview_rows),
        primary_key=None,
        editable_columns=[],
        search_column=search_column,
        search_query=search_query,
        editing_row=None,
        access_token=access_token,
        numeric_columns=_NUMERIC_COLUMNS,
        pending_change_count=0,
        preview_mode=True,
        preview_commit_id=version_id,
    )

@histology_mapping_bp.route("/dashboard/histology-code/create", methods=["POST"])
@login_required
@histology_key_required
def create_histology_code_mapping():
    columns = _get_mapping_columns()
    editable_columns = _editable_columns(columns)
    try:
        values = _validated_mapping_values(editable_columns,[request.form.get(column, "") for column in editable_columns],)
    except ValueError as error:
        flash(str(error), "danger")
        return _return_to_mapping()
    _insert_mapping_rows(columns, [values])
    flash("組織型態資料已新增。", "success")
    return _return_to_mapping()

@histology_mapping_bp.route("/dashboard/histology-code/import", methods=["POST"])
@login_required
@histology_key_required
def import_histology_code_mapping():
    uploaded_file = request.files.get("import_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("請選擇組織型態表 Excel 檔案。", "danger")
        return _return_to_mapping()
    if not uploaded_file.filename.lower().endswith(".xlsx"):
        flash("僅接受 .xlsx 格式的組織型態表。", "danger")
        return _return_to_mapping()

    try:
        rows = _read_import_rows(uploaded_file)
    except (ValueError, OSError, KeyError, BadZipFile, InvalidFileException) as error:
        flash(str(error), "danger")
        return _return_to_mapping()

    columns = _get_mapping_columns()
    editable_columns = _editable_columns(columns)
    if tuple(editable_columns) != _IMPORT_COLUMNS:
        flash("資料庫欄位與組織型態表範本不一致，無法匯入。", "danger")
        return _return_to_mapping()

    try:
        rows = [_validated_mapping_values(editable_columns, row) for row in rows]
    except ValueError as error:
        flash(str(error), "danger")
        return _return_to_mapping()

    _insert_mapping_rows(columns, rows)
    flash(f"已新增 {len(rows)} 筆組織型態資料。", "success")
    return _return_to_mapping()


@histology_mapping_bp.route("/dashboard/histology-code/<int:histcode_id>/update", methods=["POST"])
@login_required
@histology_key_required
def update_histology_code_mapping(histcode_id):
    columns = _get_mapping_columns()
    primary_key = _primary_key(columns)
    editable_columns = _editable_columns(columns)
    try:
        values = _validated_mapping_values(editable_columns,[request.form.get(column, "") for column in editable_columns],)
    except ValueError as error:
        flash(str(error), "danger")
        return _return_to_mapping()
    assignments = ", ".join(f"{_quote_identifier(column)} = ?" for column in editable_columns)
    conn = get_conn()
    cursor = conn.cursor()
    fields = ", ".join(_quote_identifier(column) for column in editable_columns)
    cursor.execute(f"SELECT {fields} FROM dbo.histology_code_mapping WHERE {_quote_identifier(primary_key)} = ?", histcode_id)
    previous = cursor.fetchone()
    cursor.execute(f"UPDATE dbo.histology_code_mapping SET {assignments} WHERE {_quote_identifier(primary_key)} = ?",*values,histcode_id,)
    record_staging_change(cursor, session["userid"], histcode_id, "Update", dict(zip(editable_columns, previous)), dict(zip(editable_columns, values)))
    conn.commit()
    conn.close()
    flash("組織型態資料已更新。", "success")
    return _return_to_mapping()


@histology_mapping_bp.route("/dashboard/histology-code/<int:histcode_id>/delete", methods=["POST"])
@login_required
@histology_key_required
def delete_histology_code_mapping(histcode_id):
    columns = _get_mapping_columns()
    primary_key = _primary_key(columns)

    conn = get_conn()
    cursor = conn.cursor()
    editable_columns = _editable_columns(columns)
    fields = ", ".join(_quote_identifier(column) for column in editable_columns)
    cursor.execute(f"SELECT {fields} FROM dbo.histology_code_mapping WHERE {_quote_identifier(primary_key)} = ?", histcode_id)
    previous = cursor.fetchone()
    cursor.execute(f"DELETE FROM dbo.histology_code_mapping WHERE {_quote_identifier(primary_key)} = ?",histcode_id,)
    record_staging_change(cursor, session["userid"], histcode_id, "Delete", before=dict(zip(editable_columns, previous)))
    conn.commit()
    conn.close()
    flash("組織型態資料已刪除。", "success")
    return _return_to_mapping()

@histology_mapping_bp.route("/dashboard/histology-code/delete-selected", methods=["POST"])
@login_required
@histology_key_required
def delete_selected_histology_code_mappings():
    histcode_ids = [value for value in request.form.getlist("histcode_ids") if value.isdigit()]
    if not histcode_ids:
        flash("請先勾選要刪除的組織型態資料。", "danger")
        return _return_to_mapping()

    columns = _get_mapping_columns()
    primary_key = _primary_key(columns)

    placeholders = ", ".join("?" for _ in histcode_ids)
    conn = get_conn()
    cursor = conn.cursor()
    editable_columns = _editable_columns(columns)
    fields = ", ".join(_quote_identifier(column) for column in editable_columns)
    cursor.execute(f"SELECT {_quote_identifier(primary_key)}, {fields} FROM dbo.histology_code_mapping WHERE {_quote_identifier(primary_key)} IN ({placeholders})", *histcode_ids)
    previous_rows = cursor.fetchall()
    cursor.execute(f"DELETE FROM dbo.histology_code_mapping WHERE {_quote_identifier(primary_key)} IN ({placeholders})",*histcode_ids,)
    for row in previous_rows:
        record_staging_change(cursor, session["userid"], row[0], "Delete", before=dict(zip(editable_columns, row[1:])))
    conn.commit()
    conn.close()
    flash(f"已刪除 {len(histcode_ids)} 筆組織型態資料。", "success")
    return _return_to_mapping()