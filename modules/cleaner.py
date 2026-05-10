import pandas as pd
import os
import re
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Font,Alignment
from modules.clean_pipeline.fmt_42 import RULES as RULES_42
from modules.clean_pipeline.fmt_45 import RULES as RULES_45
from modules.clean_pipeline.fmt_50 import RULES as RULES_50
from modules.clean_pipeline.fmt_114 import RULES as RULES_114
from modules.clean_pipeline.fmt_115 import RULES as RULES_115
from modules.clean_pipeline.fmt_129 import RULES as RULES_129
from modules.clean_pipeline.validate import check_error_type,validate_date_rules
from modules.field_mapping import field_mapping
from datetime import datetime

FORMAT_RULES_MAP = {
    "fmt_42": RULES_42,
    "fmt_45": RULES_45,
    "fmt_50": RULES_50,
    "fmt_114": RULES_114,
    "fmt_115": RULES_115,
    "fmt_129": RULES_129,
}

def annotation(row_errors):
    codes = []
    if (row_errors == "missing").any():
        codes.append("A:遺漏值")
    if (row_errors == "format").any():
        codes.append("B:格式不符")
    if (row_errors == "dateformat").any():
        codes.append("C:邏輯錯誤")
    if not codes:
        return "D:完全正確的資訊"
    return " ".join(codes)

def cleanValidate(input_file,output_file,report_file,fmt,version,Revision_Date):
    rules = FORMAT_RULES_MAP[fmt]
    file_ext = os.path.splitext(input_file)[1].lower()
    
    if file_ext == '.csv':
        try:
            df = pd.read_csv(input_file, dtype=str, encoding='utf-8-sig')
        except UnicodeDecodeError:
            df = pd.read_csv(input_file, dtype=str, encoding='cp950')
    else:
        df = pd.read_excel(input_file,dtype=str)

    alias_mapping, _ = field_mapping('中文欄位名稱')
    error_mask = pd.DataFrame("", index=df.index, columns=df.columns) 
    clean_alias_mapping = {re.sub(r'\s+', '', str(k)): v for k, v in alias_mapping.items()}
    
    for col in df.columns:
        raw_col = str(col)
        clean_col = re.sub(r'\s+', '', raw_col)
        rule_name = clean_alias_mapping.get(clean_col)
        rule = None
        if rule_name and rule_name in rules:
            rule = rules[rule_name]
        if rule:
            error_mask[col] = df[col].apply(lambda x: check_error_type(x, rule))

    for idx, row in df.iterrows():
        bad_cols, msgs = validate_date_rules(row, alias_mapping)
        if bad_cols:
            for col in bad_cols:
                if col in error_mask.columns and error_mask.at[idx, col] == "":
                    error_mask.at[idx, col] = "dateformat"

    df['_has_error'] = (error_mask != "").any(axis=1)
    df['錯誤註記說明(A:遺漏值 B:格式不符 C:邏輯錯誤 D:完全正確)'] = error_mask.apply(annotation, axis=1)
    
    # 將錯誤資料排到最上面
    sorted_df = df.sort_values(by='_has_error', ascending=False, kind='mergesort')
    sorted_mask = error_mask.loc[sorted_df.index]
    sorted_df = sorted_df.drop(columns=['_has_error'])
    sorted_df.to_excel(output_file, index=False, engine='openpyxl')

    # 錯誤標記
    wb = load_workbook(output_file)
    ws = wb.active
    fill_missing = PatternFill("solid", fgColor="FFE153")    # 黃底=遺漏
    fill_format = PatternFill("solid", fgColor="FF9797")     # 紅底=格式錯
    fill_dateformat = PatternFill("solid", fgColor="00FFFF") # 藍底=日期邏輯錯
    
    for r_idx in range(len(sorted_df)):
        for c_idx in range(len(sorted_df.columns)):
            cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
            cell.number_format = '@' 
            
            # 只有在 error_mask 範圍內的欄位才進行顏色標記
            if c_idx < len(sorted_mask.columns):
                err_type = sorted_mask.iloc[r_idx, c_idx]
                if err_type == "missing":
                    cell.fill = fill_missing
                elif err_type == "format":
                    cell.fill = fill_format
                elif err_type == "dateformat": 
                    cell.fill = fill_dateformat

    wb.save(output_file)
    wb.close()

    total_count = len(df)
    missing_rows = (error_mask == "missing").any(axis=1).sum()
    format_rows = (error_mask == "format").any(axis=1).sum()
    logic_rows = (error_mask == "dateformat").any(axis=1).sum()

    completeness = (total_count-missing_rows)/total_count
    correctness = (total_count-format_rows)/total_count
    consistency = (total_count-logic_rows)/total_count
    quality_score = (completeness*0.2+correctness*0.5+consistency*0.3)*100
    
    error_mask_bool = (error_mask != "")
    error_count = error_mask_bool.any(axis=1).sum()
    overall_accuracy = (total_count - error_count) / total_count
    missing_cell_count = (error_mask == "missing").sum().sum()
    format_cell_count = (error_mask == "format").sum().sum()
    dateformat_cell_count = (error_mask == "dateformat").sum().sum()
    
    stats = {
        'total': total_count,
        'error_rows': error_count,
        'overall_accuracy': overall_accuracy,
        'missing_rows': missing_rows,
        'format_rows': format_rows,
        'logic_rows': logic_rows,
        'completeness': completeness,
        'correctness': correctness,
        'consistency': consistency,
        'quality_score': quality_score,
        'missing_cells': missing_cell_count,
        'format_cells': format_cell_count,
        'logic_cells': dateformat_cell_count
    }

    wb_report = Workbook()
    ws_report = wb_report.active
    ws_report.title = "清洗結果"
    title_text = f"癌症登記資料清洗結果"
    ws_report.merge_cells('A1:C1')
    cell_title = ws_report['A1']
    cell_title.value = title_text
    cell_title.font = Font(size=16, bold=True)
    cell_title.alignment = Alignment(horizontal='center')
    
    report_data = [
        ["匯入日期：", datetime.now().strftime("%Y/%m/%d")],
        ["資料格式：", f"{fmt}({version}：{Revision_Date})"],
        [],
        ["資料總件數：",stats['total'],"件"],
        [],
        ["資料錯誤件數"],
        ["A:遺漏值件數：",stats['missing_rows'],"件"],
        ["B:格式不符件數：",stats['format_rows'],"件"],
        ["C:邏輯錯誤件數：",stats['logic_rows'],"件"],
        [],
        ["資料完整率：",f"{stats['completeness']:.2%}","註：完整率=(資料總件數-遺漏值件數)/資料總件數"],
        ["資料正確率：",f"{stats['correctness']:.2%}","註：正確率=(資料總件數-格式不符件數)/資料總件數"],
        ["資料一致率：",f"{stats['consistency']:.2%}","註：一致率=(資料總件數-邏輯錯誤件數)/資料總件數"],
        [],
        ["資料品質(DQI)",f"{stats['quality_score']:.2f}%","DQI(Data Quality Index)=完整率*0.2+正確率*0.5+一致率*0.3"],
        [],
        ["檢核者核章："]
    ]
    
    for r_idx, row in enumerate(report_data, start=3):
        if len(row) > 0:
            ws_report.cell(row=r_idx, column=1).value = row[0]
        if len(row) > 1:
            cell_b = ws_report.cell(row=r_idx, column=2)
            cell_b.value = row[1]
            if row[0] == "資料品質(DQI)":
                cell_b.font = Font(size=14, bold=True, color="FF0000")
                cell_b.alignment = Alignment(horizontal='center')
            else:
                cell_b.font = Font(color="0040ff") 
                cell_b.alignment = Alignment(horizontal='center')
        if len(row) > 2:
            ws_report.cell(row=r_idx, column=3).value = row[2]
    
    ws_report.column_dimensions['A'].width = 25
    ws_report.column_dimensions['B'].width = 20
    ws_report.column_dimensions['C'].width = 10 
    wb_report.save(report_file)
    wb_report.close()
    print(f"Data Clean Report file: {report_file}")
    return stats, alias_mapping, sorted_df, sorted_mask

if __name__ == "__main__": 
    input_file = 'data/HSP_TEST.xlsx'
    sheet_name = '1150318虛擬V1(給虎科)'
    output_file = f"Validate_{sheet_name}.xlsx"
    fmt = "fmt_42"
    version = "2025v1"
    Revision_Date = "2017/12/04"
    stats, _, _, _ = cleanValidate(input_file,f"{fmt}_data.xlsx",f"{fmt}_Report.xlsx",fmt,version,Revision_Date)