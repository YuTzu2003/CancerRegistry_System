import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from modules.rules import RULES
from modules.validate import validate_cell
from field_mapping import field_mapping,detect_system

input_file = 'data/20260318測試.xlsx'
sheet_name = '1150318虛擬V1(給虎科)'
output_file = f"validate_{sheet_name}.xlsx"

df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)

# system_name, _ = detect_system(df.columns)
# print(f"The data for {system_name}")

alias_mapping, _ = field_mapping('中文欄位名稱')

# 錯誤紀錄
error_mask = pd.DataFrame(False, index=df.index, columns=df.columns)
for col in df.columns:
    clean_col = str(col).strip()
    rule_name = alias_mapping.get(clean_col)

    if rule_name and rule_name in RULES:
        rule = RULES[rule_name]
        error_mask[col] = df[col].apply(lambda x: not validate_cell(x, rule))
    elif clean_col in RULES:
        rule = RULES[clean_col]
        error_mask[col] = df[col].apply(lambda x: not validate_cell(x, rule))


# 將錯誤資料排到最上面
df['_has_error'] = error_mask.any(axis=1)
sorted_df = df.sort_values(by='_has_error', ascending=False, kind='mergesort')
sorted_mask = error_mask.loc[sorted_df.index]
sorted_df = sorted_df.drop(columns=['_has_error'])
sorted_df.to_excel(output_file, index=False, engine='openpyxl')


# 錯誤標記
wb = load_workbook(output_file)
ws = wb.active
fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for r_idx in range(len(sorted_df)):
    for c_idx in range(len(sorted_df.columns)):
        if sorted_mask.iloc[r_idx, c_idx]:
    
            ws.cell(row=r_idx + 2, column=c_idx + 1).fill = fill

wb.save(output_file)
error_count = error_mask.any(axis=1).sum()

print(f"error count: {error_count}")
print(f"save file:{output_file}")